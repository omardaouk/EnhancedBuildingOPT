"""
Created on Wed July 13 14:22:00 2019

@author: Omar Daouk
"""
from __future__ import division
import Omar_MA_Code as opti_model


from openpyxl import load_workbook
from openpyxl.styles import (Font, Border, Side, Alignment)
from openpyxl import Workbook
from openpyxl.chart import (AreaChart, ScatterChart, Reference, Series)
import datetime





def run_CO2_vs_Storage_typ_wochen(number_simulations, name_of_file, name_of_x, name_of_y1, name_of_y2, storage_max,
                                  Variable, enev_restrictions, pv_scenario, status_quo_with_storage,
                                  folder="results"):

    print('Creating Output Workbook...')
    wb2 = Workbook()
    ws2 = wb2.create_sheet('Results', 0)
    ws2.cell(row=1, column=2).value = 'Simulation #'
    ws2.cell(row=1, column=3).value = 'Battery Capacity'
    ws2.cell(row=1, column=4).value = 'TES Capacity'
    ws2.cell(row=1, column=5).value = 'Total Storage Capacity'
    ws2.cell(row=1, column=6).value = 'Minimal Emissions'
    ws2.cell(row=1, column=7).value = 'Emi-gas'
    ws2.cell(row=1, column=8).value = 'Emi-pv'
    ws2.cell(row=1, column=9).value = 'Emi-grid'
    ws2.cell(row=1, column=10).value = 'Emi-lca'
    ws2.cell(row=1, column=11).value = 'Costs'
    ws2.cell(row=1, column=12).value = '# Battery Equivalent Full Cycles'
    ws2.cell(row=1, column=13).value = 'Area PV'
    ws2.cell(row=1, column=14).value = 'Area STC'
    ws2.cell(row=1, column=15).value = 'LCA PV'
    ws2.cell(row=1, column=16).value = 'LCA BATT'
    ws2.cell(row=1, column=17).value = 'LCA STC'
    ws2.cell(row=1, column=18).value = 'LCA TES'

    if pv_scenario:
        tag = "pv"
        filename_start_values = "start_values_pv.csv"
    else:
        if enev_restrictions:
            tag = "enev_restrictions"
            filename_start_values = "start_values_enev.csv"
        else:
            tag = "no_restrictions"
            filename_start_values = "start_values_without_enev.csv"

    emissions_max = 1000  # ton CO2 per year
    filename_min_costs = folder + "/" + tag + str(0) + ".pkl"

    options = {"filename_results": filename_min_costs,
               "enev_restrictions": enev_restrictions,
               "pv_scenario": pv_scenario,
               "status_quo_with_storage": status_quo_with_storage,
               "total_storage": 0,  # starting value for the storage
               "max_storage": storage_max,
               "Variable": Variable,
               "opt_costs": False,
               "store_start_vals": False,
               "load_start_vals": True,
               "filename_start_vals": filename_start_values,
               "envelope_runs": False}

    storage_increment = 0
    if number_simulations > 0:
        storage_increment = storage_max / number_simulations  # the larger the number of runs the "higher the resolution" ... i will have more points on the x axis

    for i in range(0, number_simulations + 1):  # i want  to also show the without storage thingy ... so i starts from 0

        print('############################' + '\n' + '\n')

        print("Running simulation number " + str(i) + " of " + str(number_simulations))

        print('############################' + '\n' + '\n')
        options["total_storage"] = 0 + i * storage_increment

        (max_costs, min_emissions, emi_gas, emi_pv, emi_grid, emi_lca, decision_variables, capacities, powers, storage_Batt ,results_of_clustering) = opti_model.optimize_MA(emissions_max, options)
        Battery_size = capacities["Battery"]
        TES_size = capacities["TES"]
        p_batt_charge = powers["Battery"]["total", "charge"]

        ws2.cell(row=2 + i, column=2).value = i
        ws2.cell(row=2 + i, column=3).value = Battery_size
        ws2.cell(row=2 + i, column=4).value = TES_size
        ws2.cell(row=2 + i, column=5).value = Battery_size + TES_size
        ws2.cell(row=2 + i, column=6).value = min_emissions
        ws2.cell(row=2 + i, column=7).value = emi_gas
        ws2.cell(row=2 + i, column=8).value = emi_pv
        ws2.cell(row=2 + i, column=9).value = emi_grid
        ws2.cell(row=2 + i, column=10).value = emi_lca
        ws2.cell(row=2 + i, column=11).value = max_costs
        ws2.cell(row=2 + i, column=13).value = capacities["PV" ] /0.125
        ws2.cell(row=2 + i, column=14).value = capacities["STC"]
        ws2.cell(row=2 + i, column=15).value = (capacities["PV" ] /0.125) * (0.304 / 20)
        ws2.cell(row=2 + i, column=16).value = (Battery_size * 243.9) / (1000 * 13.7)
        ws2.cell(row=2 + i, column=17).value = (capacities["STC"] * 104.3) / (1000 * 30)
        ws2.cell(row=2 + i, column=18).value = (capacities["TES"] * 3.60 * 200) / (1000 * 20)

        test =0
        delete = 0

        if Battery_size > 0:
            sigma_p_batt_charge = 0
            for j in range(len(p_batt_charge)):
                batt_charge_day =0
                for jj in range(len(p_batt_charge[j])):
                    batt_charge_day += p_batt_charge[j][jj]
                    test += 1
                sigma_p_batt_charge += batt_charge_day * results_of_clustering["weights"][j]
                delete += 1

            ws2.cell(row=2 + i, column=12).value = sigma_p_batt_charge /Battery_size

    for i in range(1, (ws2.max_row + 1)):
        ws2.row_dimensions[i].height = 15
    for i in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
        ws2.column_dimensions[i].width = 20

    # Additional code for Plotting the results Automatically

    # Chart1 for Emissions vs Capacity
    chart1 = ScatterChart()
    chart1.title = "Min-Emissions"
    chart1.style = 13
    chart1.x_axis.title = name_of_x
    chart1.y_axis.title = name_of_y1

    if name_of_x == "Battery capacity in kwh":
        xvalues = Reference(ws2, min_col=3, min_row=2, max_row=number_simulations + 2)
    elif name_of_x == "TES capacity in m3":
        xvalues = Reference(ws2, min_col=4, min_row=2, max_row=number_simulations + 2)
    elif name_of_x == "Total Storage Capacity (kwh/m3)":
        xvalues = Reference(ws2, min_col=5, min_row=2, max_row=number_simulations + 2)

    values = Reference(ws2, min_col=6, min_row=1, max_row=number_simulations + 2)
    series = Series(values, xvalues, title_from_data=True)
    chart1.series.append(series)
    ws2.add_chart(chart1, "Z7")
    chart1.width = 15
    chart1.height = 15
    chart1.legend.position = 'b'
    # Chart2 for Emissions vs Capacity
    chart2 = ScatterChart()
    chart2.title = "Costs"
    chart2.style = 13
    chart2.x_axis.title = name_of_x
    chart2.y_axis.title = name_of_y2

    if name_of_x == "Battery capacity in kwh":
        xvalues = Reference(ws2, min_col=3, min_row=2, max_row=number_simulations + 2)
    elif name_of_x == "TES capacity in m3":
        xvalues = Reference(ws2, min_col=4, min_row=2, max_row=number_simulations + 2)
    elif name_of_x == "Total Storage Capacity (kwh/m3)":
        xvalues = Reference(ws2, min_col=5, min_row=2, max_row=number_simulations + 2)

    values = Reference(ws2, min_col=11, min_row=1, max_row=number_simulations + 2)
    series = Series(values, xvalues, title_from_data=True)
    chart2.series.append(series)
    ws2.add_chart(chart2, "AK7")
    chart2.width = 15
    chart2.height = 15
    chart2.legend.position = 'b'

    # Chart3 for Detailed Emissions vs Capacity
    chart3 = ScatterChart()
    chart3.title = "Detailed_Emissions"
    chart3.style = 13
    chart3.x_axis.title = name_of_x
    chart3.y_axis.title = name_of_y1

    if name_of_x == "Battery capacity in kwh":
        xvalues = Reference(ws2, min_col=3, min_row=2, max_row=number_simulations + 2)
    elif name_of_x == "TES capacity in m3":
        xvalues = Reference(ws2, min_col=4, min_row=2, max_row=number_simulations + 2)
    elif name_of_x == "Total Storage Capacity (kwh/m3)":
        xvalues = Reference(ws2, min_col=5, min_row=2, max_row=number_simulations + 2)

    for place_holder in (7, 8, 9, 10):
        values = Reference(ws2, min_col=place_holder, min_row=1, max_row=number_simulations + 2)
        series = Series(values, xvalues, title_from_data=True)
        chart3.series.append(series)
    ws2.add_chart(chart3, "Z37")
    chart3.width = 15
    chart3.height = 15
    chart3.legend.position = 'b'

    counter = number_simulations + 4
    starting_column = 2
    ws2.cell(row=counter, column=starting_column).value = 1
    ws2.cell(row=number_simulations + 5, column=starting_column).value = 'Simulation #'
    ws2.cell(row=number_simulations + 5, column=starting_column + 1).value = 'Device'
    ws2.cell(row=number_simulations + 5, column=starting_column + 2).value = 'Factor'
    ws2.cell(row=number_simulations + 5, column=starting_column + 3).value = 'Value'

    for dev in ["Battery", "Battery_small", "Battery_large", "TES", "Boiler", "CHP", "PV", "HP", "EH", "STC"]:
        ws2.cell(row=counter, column=starting_column + 1).value = dev
        ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev
        if dev in ("TES", "Boiler", "CHP"):
            for i in (0, 1):
                if dev == "TES" or dev == "Boiler":
                    if i == 0:
                        ws2.cell(row=counter, column=starting_column + 1).value = dev + "_" + str(i)
                        ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev + str(i)
                        ws2.cell(row=counter, column=starting_column + 3).value = decision_variables[dev][i]
                        counter += 1

                    else:
                        ws2.cell(row=counter, column=starting_column + 1).value = dev + "_Continuous"
                        ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev + "_Continuous"
                        ws2.cell(row=counter, column=starting_column + 3).value = decision_variables[dev][i]
                        ws2.cell(row=counter + 1, column=starting_column + 1).value = dev + "_Continuous"
                        ws2.cell(row=counter + 1,
                                 column=starting_column + 2).value = "cap_design_" + dev + "_Continuous"
                        ws2.cell(row=counter + 1, column=starting_column + 3).value = capacities[dev + "_Conti"]
                        ws2.cell(row=counter + 2, column=starting_column + 1).value = dev + "Total"
                        ws2.cell(row=counter + 2, column=starting_column + 2).value = "cap_" + dev + "_Total"
                        ws2.cell(row=counter + 2, column=starting_column + 3).value = capacities[dev]
                        counter += 3

                elif dev == "CHP":
                    ws2.cell(row=counter, column=starting_column + 1).value = dev + "_" + str(i)
                    ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev + str(i)
                    ws2.cell(row=counter, column=starting_column + 3).value = decision_variables[dev][i]
                    ws2.cell(row=counter + 1, column=starting_column + 1).value = dev
                    ws2.cell(row=counter + 1, column=starting_column + 2).value = "cap_design_" + dev
                    ws2.cell(row=counter + 1, column=starting_column + 3).value = capacities[dev + str(i)]
                    counter += 1
                    if i == 1:
                        ws2.cell(row=counter + 2, column=starting_column + 1).value = dev + "Total"
                        ws2.cell(row=counter + 2, column=starting_column + 2).value = "cap_" + dev + "_Total"
                        ws2.cell(row=counter + 2, column=starting_column + 3).value = capacities[dev]
                        counter += 1
        else:
            ws2.cell(row=counter, column=starting_column + 1).value = dev
            ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev
            ws2.cell(row=counter, column=starting_column + 3).value = decision_variables[dev]
            ws2.cell(row=counter + 1, column=starting_column + 1).value = dev
            ws2.cell(row=counter + 1, column=starting_column + 2).value = "cap_design_" + dev
            ws2.cell(row=counter + 1, column=starting_column + 3).value = capacities[dev]
            counter += 2

    number_of_representative_periods = results_of_clustering["representative_days"]["T_e_raw"].shape[0]
    length_of_cluster = results_of_clustering["representative_days"]["T_e_raw"].shape[1]

    counter += 3
    column = 4

    ws2.cell(row=counter, column=3).value = 'Power Details for last Simulation'
    for dev in ["HP", "CHP", "EH", "STC", "PV", "Import", "Battery_small", "Battery_large", "Battery", ]:
        ws2.cell(row=counter, column=column).value = dev
        rows = counter
        if dev in ("Battery_small", "Battery_large", "Battery"):
            for state in ("charge", "discharge"):
                ws2.cell(row=rows, column=column).value = dev + "_" + state
                d_multiplier = -1
                for d in range(number_of_representative_periods):
                    d_multiplier += 1
                    for t in range(length_of_cluster):
                        if d == 0 and t == 0:
                            rows += 1
                        ws2.cell(row=(rows + (d_multiplier * length_of_cluster) + t), column=column).value = \
                        powers[dev]["total", state][d][t]
                column += 1
                rows = counter
            ws2.cell(row=rows, column=column).value = "storage" + "_" + dev + "in %"
            d_multiplier = -1
            for d in range(number_of_representative_periods):
                d_multiplier += 1
                for t in range(length_of_cluster):
                    if d == 0 and t == 0:
                        rows += 1
                    ws2.cell(row=(rows + (d_multiplier * length_of_cluster) + t), column=column).value = (storage_Batt[
                                                                                                              d][t] /
                                                                                                          capacities[
                                                                                                              dev]) * 100
            column += 1
            rows = counter


        else:
            d_multiplier = -1
            for d in range(number_of_representative_periods):
                d_multiplier += 1
                for t in range(length_of_cluster):
                    if d == 0 and t == 0:
                        rows += 1
                    ws2.cell(row=(rows + (d_multiplier * length_of_cluster) + t), column=column).value = powers[dev][d][
                        t]
            column += 1
            if dev in ["PV", "CHP", ]:
                for method in ("use", "sell"):
                    rows = counter
                    ws2.cell(row=counter, column=column).value = dev + "_" + method
                    d_multiplier = -1
                    for d in range(number_of_representative_periods):
                        d_multiplier += 1
                        for t in range(length_of_cluster):
                            if d == 0 and t == 0:
                                rows += 1
                            ws2.cell(row=(rows + (d_multiplier * length_of_cluster) + t), column=column).value = \
                            powers[dev, method][d][t]
                    column += 1

    for i in range(1, (ws2.max_row + 1)):  # ws2 has fewer rows than ws1 so it will work
        ws2.row_dimensions[i].height = 15
    for i in ("A", "B", "F", "G", "K", "L", "M", "N", "O", "P", "Q"):
        ws2.column_dimensions[i].width = 15
    for i in ("C", "D", "E", "H", "I", "J"):
        ws2.column_dimensions[i].width = 25

    name_of_output = name_of_file
    current_date = datetime.datetime.now()
    wb2.save(str(current_date.month) + "." + str(current_date.day) + "." + str(current_date.year) + " - " +
             str(current_date.hour) + "_" + str(current_date.minute) + "_" + "_" + name_of_output + ".xlsx")

    # Workbook to show the clustered outputs and the weights
    wb3 = Workbook()
    ws3 = wb3.create_sheet('Clustered_Results', 0)

    list_of_data_groups = ['CO2(t)', 'dhw', 'electricity', 'sun_rad_0', 'sun_rad_1', 'sun_rad_2', 'sun_rad_3',
                           'sun_rad_4', 'sun_rad_5', 'T_e_raw']

    ws3.cell(row=1, column=1).value = "Weights of the Data types"
    ws3.cell(row=1, column=2).value = "Value of the Objective function"
    ws3.cell(row=2, column=2).value = results_of_clustering["obj"]

    for i in range(len(list_of_data_groups)):
        ws3.cell(row=2 + i, column=1).value = list_of_data_groups[i] + "=" + str(
            results_of_clustering["weights_of_input"][i])

    count_columns = 2

    nn = 0
    for j in results_of_clustering["list_size_and_elements_of_cluster"]:
        count_columns += 1
        ws3.cell(row=1, column=count_columns).value = str("Cluster Number : " + str(nn))
        ws3.cell(row=2, column=count_columns).value = str("Representative Element : " + str(j[0]))
        ws3.cell(row=3, column=count_columns).value = str("Weight of Cluster : " + str(j[1][0]))
        ws3.cell(row=4, column=count_columns).value = str("Elements in Cluster" + '_' + str(nn))
        count_rows = 5
        nn += 1
        for i in j[1][1]:
            ws3.cell(row=count_rows, column=count_columns).value = i
            count_rows += 1

    for ii in range(len(results_of_clustering["representative_days"])):
        ''' len of results_of_clustering gives the number of lists ... 
            1 List for each data type and each list has n elements ... 1 element for each of the n clusters  '''
        for jj in range(len(results_of_clustering["representative_days"][list_of_data_groups[0]])):
            '''
            len of results_of_clustering[0] gives the number of sub elements in each element
            this number of sub elements corresponds to the number of clusters ...
            each sub element is dedicated to the values of this element for a given cluster
            each element represents a Data Group
            '''
            count_columns += 1
            count_rows = 2
            for zz in range(len(results_of_clustering["representative_days"][list_of_data_groups[0]][0])):
                '''
                len of results_of_clustering[0][0] gives the number of values within each sub element
                these 24 values will be the hourly values of this representative day of cluster jj
                of element ii of the group of data to be clustered
                 '''
                ws3.cell(row=1, column=count_columns).value = str(list_of_data_groups[ii]) + '_' + str(jj)
                ws3.cell(row=count_rows, column=count_columns).value = \
                results_of_clustering["representative_days"][list_of_data_groups[ii]][jj][zz]
                count_rows += 1

    for i in range(1, (ws3.max_row + 1)):  # ws2 has fewer rows than ws1 so it will work
        ws3.row_dimensions[i].height = 15
    for i in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
              "V", "W", "X", "Y", "Z"):
        ws3.column_dimensions[i].width = 25

    current_date = datetime.datetime.now()
    wb3.save(str(current_date.month) + "." + str(current_date.day) + "." + str(current_date.year) + " - " +
             str(current_date.hour) + "_" + str(current_date.minute) + "_" + "_" + "Clustered Data 2017" + ".xlsx")
    # Excel file with the detailed Clustered Data


def run_CO2_vs_Envelope_typ_wochen(name_of_file, envelope_details, bes_details, folder="results"):
    print('Creating Output Workbook...')
    wb2 = Workbook()
    ws2 = wb2.create_sheet('Results', 0)
    ws2.cell(row=2, column=2).value = 'Simulation '
    ws2.cell(row=3, column=2).value = 'Wall'
    ws2.cell(row=4, column=2).value = 'Roof'
    ws2.cell(row=5, column=2).value = 'Window'
    ws2.cell(row=6, column=2).value = 'Minimal Emissions'
    ws2.cell(row=7, column=2).value = 'Emi-gas'
    ws2.cell(row=8, column=2).value = 'Emi-pv'
    ws2.cell(row=9, column=2).value = 'Emi-grid'
    ws2.cell(row=10, column=2).value = 'Emi-lca'
    ws2.cell(row=11, column=2).value = 'Costs'
    ws2.cell(row=12, column=2).value = 'LCA Wall'
    ws2.cell(row=13, column=2).value = 'LCA Roof'
    ws2.cell(row=14, column=2).value = 'LCA Window'
    ws2.cell(row=15, column=2).value = 'Area PV'
    ws2.cell(row=16, column=2).value = 'Area STC'
    ws2.cell(row=17, column=2).value = 'LCA PV'
    ws2.cell(row=18, column=2).value = 'LCA BATT'
    ws2.cell(row=19, column=2).value = 'LCA STC'
    ws2.cell(row=20, column=2).value = 'LCA TES'
    ws2.cell(row=21, column=2).value = 'Battery Capacity'
    ws2.cell(row=22, column=2).value = 'TES Capacity'
    ws2.cell(row=23, column=2).value = '# Battery Equivalent Full Cycles'

    LCA_envelope = {}
    LCA_envelope = {"wall 0": 0, "wall 1": 266.18 / 1000, "wall 2": 471.01 / 1000, "wall 3": 487.35 / 1000,
                    "roof 0": 0, "roof 1": 126.76 / 1000, "roof 2": 250.41 / 1000, "roof 3": 330.01 / 1000,
                    "window 0": 0, "window 1": 7.6 / 1000, "window 2": 27.95 / 1000,
                    "window 3": 37.82 / 1000}  # emissions /1000 to become in tonnes ... they were in kilos originally

    emissions_max = 1000  # ton CO2 per year

    options = {
        "opt_costs": False,
        "store_start_vals": False,
        "load_start_vals": True,
        "envelope": envelope_details,
        "bes": bes_details
    }
    options["pv_scenario"] = False
    options["status_quo_with_storage"] = False
    options["enev_restrictions"] = False
    options["envelope_runs"] = True

    pv_scenario = False
    enev_restrictions = False
    if pv_scenario:
        tag = "pv"
        filename_start_values = "start_values_pv.csv"
    else:
        if enev_restrictions:
            tag = "enev_restrictions"
            filename_start_values = "start_values_enev.csv"
        else:
            tag = "no_restrictions"
            filename_start_values = "start_values_without_enev.csv"

    options["filename_start_vals"] = filename_start_values

    simulations_to_run = {
        "Walls 0": {"wall": 0, "roof": 0, "window": 0},
        "Walls 1": {"wall": 1, "roof": 0, "window": 0},
        "Walls 2": {"wall": 2, "roof": 0, "window": 0},
        "Walls 3": {"wall": 3, "roof": 0, "window": 0},
        "Roof 0": {"wall": 0, "roof": 0, "window": 0},
        "Roof 1": {"wall": 0, "roof": 1, "window": 0},
        "Roof 2": {"wall": 0, "roof": 2, "window": 0},
        "Roof 3": {"wall": 0, "roof": 3, "window": 0},
        "Windows 0": {"wall": 0, "roof": 0, "window": 0},
        "Windows 1": {"wall": 0, "roof": 0, "window": 1},
        "Windows 2": {"wall": 0, "roof": 0, "window": 2},
        "Windows 3": {"wall": 0, "roof": 0, "window": 3},
    }

    counter_for_runs = 0
    for i in simulations_to_run.keys():
        counter_for_runs += 1
        print("running simulation", str(counter_for_runs), "of", len(simulations_to_run))
        tag = i
        filename_min_costs = folder + "/" + tag + str(0) + ".pkl"
        options["filename_results"] = filename_min_costs

        print('############################' + '\n' + '\n')

        print("Running simulation:  " + str(i))
        new_bes = bes_details  # in case we need to change BES do it here
        options["envelope"] = simulations_to_run[i]
        options["bes"] = new_bes
        print("with following envelope:", options["envelope"])
        print("with following bes:", new_bes)
        print('############################' + '\n' + '\n')
        (max_costs, min_emissions, emi_gas, emi_pv, emi_grid, emi_lca, decision_variables, capacities, powers,
         storage_Batt, results_of_clustering) = opti_model.optimize_MA(emissions_max, options)
        Battery_size = capacities["Battery"]
        TES_size = capacities["TES"]
        p_batt_charge = powers["Battery"]["total", "charge"]

        ws2.cell(row=2, column=2 + counter_for_runs).value = i
        ws2.cell(row=3, column=2 + counter_for_runs).value = simulations_to_run[i]['wall']
        ws2.cell(row=4, column=2 + counter_for_runs).value = simulations_to_run[i]['roof']
        ws2.cell(row=5, column=2 + counter_for_runs).value = simulations_to_run[i]['window']
        ws2.cell(row=6, column=2 + counter_for_runs).value = min_emissions
        ws2.cell(row=7, column=2 + counter_for_runs).value = emi_gas
        ws2.cell(row=8, column=2 + counter_for_runs).value = emi_pv
        ws2.cell(row=9, column=2 + counter_for_runs).value = emi_grid
        ws2.cell(row=10, column=2 + counter_for_runs).value = emi_lca
        ws2.cell(row=11, column=2 + counter_for_runs).value = max_costs
        ws2.cell(row=12, column=2 + counter_for_runs).value = LCA_envelope["wall " + str(simulations_to_run[i][
                                                                                             'wall'])]  # Actually much simpler ... return the emission of all walls and windows since they are calculated regardless of wheter or not they are used in gurobi ... and then i can use the ones that are used here and ignore the others
        ws2.cell(row=13, column=2 + counter_for_runs).value = LCA_envelope["roof " + str(simulations_to_run[i][
                                                                                             'roof'])]  # Actually much simpler ... return the emission of all walls and windows since they are calculated regardless of wheter or not they are used in gurobi ... and then i can use the ones that are used here and ignore the others
        ws2.cell(row=14, column=2 + counter_for_runs).value = LCA_envelope["window " + str(simulations_to_run[i][
                                                                                               'window'])]  # Actually much simpler ... return the emission of all walls and windows since they are calculated regardless of wheter or not they are used in gurobi ... and then i can use the ones that are used here and ignore the others
        ws2.cell(row=15, column=2 + counter_for_runs).value = capacities["PV"] / 0.125
        ws2.cell(row=16, column=2 + counter_for_runs).value = capacities["STC"]
        ws2.cell(row=17, column=2 + counter_for_runs).value = (capacities["PV"] / 0.125) * (0.304 / 20)
        ws2.cell(row=18, column=2 + counter_for_runs).value = (Battery_size * 243.9) / (1000 * 13.7)
        ws2.cell(row=19, column=2 + counter_for_runs).value = (capacities["STC"] * 104.3) / (1000 * 30)
        ws2.cell(row=20, column=2 + counter_for_runs).value = (capacities["TES"] * 3.60 * 200) / (1000 * 20)
        ws2.cell(row=21, column=2 + counter_for_runs).value = Battery_size
        ws2.cell(row=22, column=2 + counter_for_runs).value = TES_size

        test = 0
        delete = 0

        if Battery_size > 0:
            sigma_p_batt_charge = 0
            for j in range(len(p_batt_charge)):
                batt_charge_day = 0
                for jj in range(len(p_batt_charge[j])):
                    batt_charge_day += p_batt_charge[j][jj]
                    test += 1
                sigma_p_batt_charge += batt_charge_day * results_of_clustering["weights"][j]
                delete += 1

            ws2.cell(row=23, column=2 + counter_for_runs).value = sigma_p_batt_charge / Battery_size


    for i in range(1, (ws2.max_row + 1)):
        ws2.row_dimensions[i].height = 15
    for i in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
        ws2.column_dimensions[i].width = 20

    for i in range(1, (ws2.max_row + 1)):  # ws2 has fewer rows than ws1 so it will work
        ws2.row_dimensions[i].height = 15
    for i in ("A", "B", "F", "G", "K", "L", "M", "N", "O", "P", "Q"):
        ws2.column_dimensions[i].width = 15
    for i in ("C", "D", "E", "H", "I", "J"):
        ws2.column_dimensions[i].width = 25

    name_of_output = name_of_file
    current_date = datetime.datetime.now()
    wb2.save(str(current_date.month) + "." + str(current_date.day) + "." + str(current_date.year) + " - " +
             str(current_date.hour) + "_" + str(current_date.minute) + "_" + "_" + name_of_output + ".xlsx")

    """NEW CLUSTER DATA"""
    # Workbook to show the clustered outputs and the weights
    wb3 = Workbook()
    ws3 = wb3.create_sheet('Clustered_Results', 0)

    list_of_data_groups = ['CO2(t)', 'dhw', 'electricity', 'sun_rad_0', 'sun_rad_1', 'sun_rad_2', 'sun_rad_3',
                           'sun_rad_4', 'sun_rad_5', 'T_e_raw']

    ws3.cell(row=1, column=1).value = "Weights of the Data types"
    ws3.cell(row=1, column=2).value = "Value of the Objective function"
    ws3.cell(row=2, column=2).value = results_of_clustering["obj"]

    for i in range(len(list_of_data_groups)):
        ws3.cell(row=2 + i, column=1).value = list_of_data_groups[i] + "=" + str(
            results_of_clustering["weights_of_input"][i])

    count_columns = 2

    """ADDED NEXT LINE HERE TODAY 03.09.2019"""

    nn = 0
    for j in results_of_clustering["list_size_and_elements_of_cluster"]:
        count_columns += 1
        ws3.cell(row=1, column=count_columns).value = str("Cluster Number : " + str(nn))
        ws3.cell(row=2, column=count_columns).value = str("Representative Element : " + str(j[0]))
        ws3.cell(row=3, column=count_columns).value = str("Weight of Cluster : " + str(j[1][0]))
        ws3.cell(row=4, column=count_columns).value = str("Elements in Cluster" + '_' + str(nn))
        count_rows = 5
        nn += 1
        for i in j[1][1]:
            ws3.cell(row=count_rows, column=count_columns).value = i
            count_rows += 1

    # count_columns = 4

    for ii in range(len(results_of_clustering["representative_days"])):
        ''' len of results_of_clustering gives the number of lists ... 
            1 List for each data type and each list has n elements ... 1 element for each of the n clusters  '''
        for jj in range(len(results_of_clustering["representative_days"][list_of_data_groups[0]])):
            '''
            len of results_of_clustering[0] gives the number of sub elements in each element
            this number of sub elements corresponds to the number of clusters ...
            each sub element is dedicated to the values of this element for a given cluster
            each element represents a Data Group
            '''
            count_columns += 1
            count_rows = 2
            for zz in range(len(results_of_clustering["representative_days"][list_of_data_groups[0]][0])):
                '''
                len of results_of_clustering[0][0] gives the number of values within each sub element
                these 24 values will be the hourly values of this representative day of cluster jj
                of element ii of the group of data to be clustered
                 '''
                ws3.cell(row=1, column=count_columns).value = str(list_of_data_groups[ii]) + '_' + str(jj)
                ws3.cell(row=count_rows, column=count_columns).value = \
                    results_of_clustering["representative_days"][list_of_data_groups[ii]][jj][zz]
                count_rows += 1

    for i in range(1, (ws3.max_row + 1)):  # ws2 has fewer rows than ws1 so it will work
        ws3.row_dimensions[i].height = 15
    for i in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
              "V", "W", "X", "Y", "Z"):
        ws3.column_dimensions[i].width = 25

    current_date = datetime.datetime.now()
    wb3.save(str(current_date.month) + "." + str(current_date.day) + "." + str(current_date.year) + " - " +
             str(current_date.hour) + "_" + str(current_date.minute) + "_" + "_" + "Clustered Data 2017" + ".xlsx")
    # Excel file with the detailed Clustered Data


def run_multi_obj(name_of_output, number_simulations, enev_restrictions=True, pv_scenario=False,
                  status_quo_with_storage=False, folder="results"):
    # Filename definitions
    if pv_scenario:
        tag = "pv"
        filename_start_values = "start_values_pv.csv"
    else:
        if enev_restrictions:
            tag = "enev_restrictions"
            filename_start_values = "start_values_enev.csv"
        else:
            tag = "no_restrictions"
            filename_start_values = "start_values_without_enev.csv"

    # Compute limits (min costs, min emissions)
    print(
        '################' + '\n' + '\n' + 'Running the First optimization (1.optimize) to compute minimum cost and maximum emissions')
    emissions_max = 1000  # ton CO2 per year
    # Minimize costs
    filename_min_costs = folder + "/" + tag + str(0) + ".pkl"

    options = {"filename_results": filename_min_costs,
               "enev_restrictions": enev_restrictions,
               "pv_scenario": pv_scenario,
               "status_quo_with_storage": status_quo_with_storage,
               "opt_costs": True,
               "store_start_vals": False,
               "load_start_vals": False,
               "filename_start_vals": filename_start_values,
               "envelope_runs": False}

    (min_costs1, max_emissions1, emi_gas1, emi_pv1, emi_grid1, emi_lca1, decision_variables_1, capacities_1, powers_1,
     storage_Batt_1, clustered_data) = opti_model.optimize_MA(emissions_max, options)

    results_of_first = (
        min_costs1, max_emissions1, emi_gas1, emi_pv1, emi_grid1, emi_lca1, decision_variables_1, capacities_1,
        powers_1,
        storage_Batt_1)

    # Minimize emissions (lexicographic optimization)
    print(
        '################' + '\n' + '\n' + 'Running the Second optimization (2.optimize) to compute minimum emissions and maximum costs')
    filename_min_emissions = folder + "/" + tag + str(number_simulations + 1) + ".pkl"
    options["opt_costs"] = False
    options["store_start_vals"] = True
    options["filename_results"] = filename_min_emissions
    (max_costs2, min_emissions2, emi_gas2, emi_pv2, emi_grid2, emi_lca2, decision_variables_2, capacities_2, powers_2,
     storage_Batt_2, clustered_data) = opti_model.optimize_MA(emissions_max, options)

    results_of_second = (
        max_costs2, min_emissions2, emi_gas2, emi_pv2, emi_grid2, emi_lca2, decision_variables_2, capacities_2,
        powers_2,
        storage_Batt_2)

    # Second optimization to minimize the costs at minimal emissions
    # print('################' + '\n' + '\n' +'Running the Third optimization (3.optimize) to compute minimum emissions and minimum costs')
    # options["opt_costs"] = True
    # options["store_start_vals"] = True
    # options["load_start_vals"] = True
    # options["filename_results"] = filename_min_emissions
    # (max_costs3, min_emissions3, emi_gas3, emi_pv3, emi_grid3, emi_lca3, decision_variables_3, capacities_3, powers_3, storage_Batt_3 ) = opti_model.optimize(min_emissions, options)
    """The second optimization is commented out because for some simulations it gives a problem since
    minimizing the costs for the set emissions is not possible"""


    # Run multiple simulations
    print(
        '################' + '\n' + '\n' + 'Running the Fourth optimization (4.optimize) to compute the multiple simulations of possible solutions and get the pareto front')
    options["opt_costs"] = True
    options["store_start_vals"] = False
    options["load_start_vals"] = True
    prev_emissions = max_emissions1
    results_of_fourth = {}
    for i in range(1, 1 + number_simulations):
        # Emissions limit is the minimum of:
        # 1. linear interpolation between max_emissions and min_emissions
        # 2. previous iteration's emissions * (1-eps)
        limit_emissions = min(max_emissions1 - (max_emissions1 - min_emissions2) * i / (number_simulations + 1),
                              prev_emissions * 0.999)
        print("################")
        print("################")
        print(str(1 + number_simulations - i) + " Simulations left")
        print("################")
        print("################")

        options["filename_results"] = folder + "/" + tag + str(i) + ".pkl"
        (costs, prev_emissions, emi_gas4, emi_pv4, emi_grid4, emi_lca4, decision_variables_4, capacities_4, powers_4,
         storage_Batt_4, clustered_data) \
            = opti_model.optimize_MA(limit_emissions, options)

        results_of_fourth[i] = (costs, prev_emissions, emi_gas4, emi_pv4, emi_grid4, emi_lca4, decision_variables_4,
                                capacities_4, powers_4, storage_Batt_4)

    ####

    # Plotting the results

    print('Creating Output Workbook...')
    wb2 = Workbook()
    ws2 = wb2.create_sheet('Results', 0)
    ws2.cell(row=1, column=2).value = 'Simulation #'
    ws2.cell(row=1, column=3).value = 'Costs'
    ws2.cell(row=1, column=4).value = 'Emissions'
    ws2.cell(row=1, column=5).value = 'Emi-gas'
    ws2.cell(row=1, column=6).value = 'Emi-pv'
    ws2.cell(row=1, column=7).value = 'Emi-grid'
    ws2.cell(row=1, column=8).value = 'Emi-lca'

    counter = 2
    for all in [results_of_first, results_of_fourth, results_of_second]:
        if all == results_of_fourth:
            for ii in range(1, 1 + number_simulations):
                ws2.cell(row=counter, column=2).value = ii + 2
                ws2.cell(row=counter, column=3).value = all[ii][0]
                ws2.cell(row=counter, column=4).value = all[ii][1]
                ws2.cell(row=counter, column=5).value = all[ii][2]
                ws2.cell(row=counter, column=6).value = all[ii][3]
                ws2.cell(row=counter, column=7).value = all[ii][4]
                ws2.cell(row=counter, column=8).value = all[ii][5]
                counter += 1
        else:
            if all == results_of_first:
                ws2.cell(row=counter, column=2).value = 1
            elif all == results_of_second:
                ws2.cell(row=counter, column=2).value = 2

            ws2.cell(row=counter, column=3).value = all[0]
            ws2.cell(row=counter, column=4).value = all[1]
            ws2.cell(row=counter, column=5).value = all[2]
            ws2.cell(row=counter, column=6).value = all[3]
            ws2.cell(row=counter, column=7).value = all[4]
            ws2.cell(row=counter, column=8).value = all[5]
            counter += 1

    chart1 = ScatterChart()
    chart1.title = "Min-Emissions"
    chart1.style = 13
    chart1.x_axis.title = 'Costs in 1000 Euro per year'
    chart1.y_axis.title = 'Emissions in t CO2 per year'
    xvalues = Reference(ws2, min_col=3, min_row=2, max_row=number_simulations + 3)
    values = Reference(ws2, min_col=4, min_row=1, max_row=number_simulations + 3)
    series = Series(values, xvalues, title_from_data=True)
    chart1.series.append(series)
    ws2.add_chart(chart1, "M2")
    chart1.width = 15
    chart1.height = 15
    chart1.legend.position = 'b'

    ####

    # Device specific results

    # for all in [results_of_first, results_of_fourth, results_of_second]:
    for simulation_results in [results_of_first, results_of_second]:
        counter = number_simulations + 6
        if simulation_results == results_of_first:
            starting_column = 2
            ws2.cell(row=counter, column=starting_column).value = 1


        else:
            starting_column = 7
            ws2.cell(row=counter, column=starting_column).value = 2

        ws2.cell(row=number_simulations + 5, column=starting_column).value = 'Simulation #'
        ws2.cell(row=number_simulations + 5, column=starting_column + 1).value = 'Device'
        ws2.cell(row=number_simulations + 5, column=starting_column + 2).value = 'Factor'
        ws2.cell(row=number_simulations + 5, column=starting_column + 3).value = 'Value'

        for dev in ["Battery", "Battery_small", "Battery_large", "TES", "Boiler", "CHP", "PV", "HP", "EH", "STC"]:
            ws2.cell(row=counter, column=starting_column + 1).value = dev
            ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev
            if dev in ("TES", "Boiler", "CHP"):
                for i in (0, 1):
                    if dev == "TES" or dev == "Boiler":
                        if i == 0:
                            ws2.cell(row=counter, column=starting_column + 1).value = dev + "_" + str(i)
                            ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev + str(i)
                            ws2.cell(row=counter, column=starting_column + 3).value = simulation_results[6][dev][i]
                            counter += 1

                        else:
                            ws2.cell(row=counter, column=starting_column + 1).value = dev + "_Continuous"
                            ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev + "_Continuous"
                            ws2.cell(row=counter, column=starting_column + 3).value = simulation_results[6][dev][i]
                            ws2.cell(row=counter + 1, column=starting_column + 1).value = dev + "_Continuous"
                            ws2.cell(row=counter + 1,
                                     column=starting_column + 2).value = "cap_design_" + dev + "_Continuous"
                            ws2.cell(row=counter + 1, column=starting_column + 3).value = simulation_results[7][
                                dev + "_Conti"]
                            ws2.cell(row=counter + 2, column=starting_column + 1).value = dev + "Total"
                            ws2.cell(row=counter + 2, column=starting_column + 2).value = "cap_" + dev + "_Total"
                            ws2.cell(row=counter + 2, column=starting_column + 3).value = simulation_results[7][dev]
                            counter += 3

                    elif dev == "CHP":
                        ws2.cell(row=counter, column=starting_column + 1).value = dev + "_" + str(i)
                        ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev + str(i)
                        ws2.cell(row=counter, column=starting_column + 3).value = simulation_results[6][dev][i]
                        ws2.cell(row=counter + 1, column=starting_column + 1).value = dev
                        ws2.cell(row=counter + 1, column=starting_column + 2).value = "cap_design_" + dev
                        ws2.cell(row=counter + 1, column=starting_column + 3).value = simulation_results[7][
                            dev + str(i)]
                        counter += 1
                        if i == 1:
                            ws2.cell(row=counter + 2, column=starting_column + 1).value = dev + "Total"
                            ws2.cell(row=counter + 2, column=starting_column + 2).value = "cap_" + dev + "_Total"
                            ws2.cell(row=counter + 2, column=starting_column + 3).value = simulation_results[7][dev]
                            counter += 1
            else:
                ws2.cell(row=counter, column=starting_column + 1).value = dev
                ws2.cell(row=counter, column=starting_column + 2).value = "x_" + dev
                ws2.cell(row=counter, column=starting_column + 3).value = simulation_results[6][dev]
                ws2.cell(row=counter + 1, column=starting_column + 1).value = dev
                ws2.cell(row=counter + 1, column=starting_column + 2).value = "cap_design_" + dev
                ws2.cell(row=counter + 1, column=starting_column + 3).value = simulation_results[7][dev]
                counter += 2

    counter += 3
    column = 4

    ws2.cell(row=counter, column=3).value = 'Power Details for the 2nd Simulation'
    for dev in ["HP", "CHP", "EH", "STC", "PV", "Import", "Battery", "Battery_small", "Battery_large"]:
        ws2.cell(row=counter, column=column).value = dev
        rows = counter
        if dev == "Battery" or dev == "Battery_small" or dev == "Battery_large":
            for state in ("charge", "discharge"):
                ws2.cell(row=rows, column=column).value = dev + "_" + state
                d_multiplier = -1
                for d in range(len(results_of_first[8][dev]["total", state])):
                    d_multiplier += 1
                    for t in range(len(results_of_first[8][dev]["total", state][d])):
                        if d == 0 and t == 0:
                            rows += 1
                        ws2.cell(row=(rows + (d_multiplier * len(results_of_first[8][dev]["total", state][d])) + t),
                                 column=column).value = results_of_first[8][dev]["total", state][d][t]
                        ws2.cell(row=(rows + (d_multiplier * len(results_of_first[8][dev]["total", state][d])) + t),
                                 column=3).value = (str(d) + "_" + str(t))
                column += 1
                rows = counter
        else:
            d_multiplier = -1
            for d in range(len(results_of_first[8][dev])):
                d_multiplier += 1
                for t in range(len(results_of_first[8][dev][d])):
                    if d == 0 and t == 0:
                        rows += 1
                    ws2.cell(row=(rows + (d_multiplier * len(results_of_first[8][dev][d])) + t), column=column).value = \
                    results_of_first[8][dev][d][t]
            column += 1
            if dev in ["PV", "CHP", ]:
                for method in ("use", "sell"):
                    rows = counter
                    ws2.cell(row=counter, column=column).value = dev + "_" + method
                    d_multiplier = -1
                    for d in range(len(results_of_first[8][dev, method])):
                        d_multiplier += 1
                        for t in range(len(results_of_first[8][dev, method][d])):
                            if d == 0 and t == 0:
                                rows += 1
                            ws2.cell(row=(rows + (d_multiplier * len(results_of_first[8][dev, method][d])) + t),
                                     column=column).value = \
                                results_of_first[8][dev, method][d][t]
                    column += 1

    column += 2

    ws2.cell(row=counter, column=21).value = 'Power Details for the 1st Simulation'
    for dev in ["HP", "CHP", "EH", "STC", "PV", "Import", "Battery", "Battery_small", "Battery_large"]:
        ws2.cell(row=counter, column=column).value = dev
        rows = counter
        if dev == "Battery" or dev == "Battery_small" or dev == "Battery_large":
            for state in ("charge", "discharge"):
                ws2.cell(row=rows, column=column).value = dev + "_" + state
                d_multiplier = -1
                for d in range(len(results_of_second[8][dev]["total", state])):
                    d_multiplier += 1
                    for t in range(len(results_of_second[8][dev]["total", state][d])):
                        if d == 0 and t == 0:
                            rows += 1
                        ws2.cell(row=(rows + (d_multiplier * len(results_of_second[8][dev]["total", state][d])) + t),
                                 column=column).value = results_of_second[8][dev]["total", state][d][t]
                        ws2.cell(row=(rows + (d_multiplier * len(results_of_second[8][dev]["total", state][d])) + t),
                                 column=3).value = (str(d) + "_" + str(t))
                column += 1
                rows = counter
        else:
            d_multiplier = -1
            for d in range(len(results_of_second[8][dev])):
                d_multiplier += 1
                for t in range(len(results_of_second[8][dev][d])):
                    if d == 0 and t == 0:
                        rows += 1
                    ws2.cell(row=(rows + (d_multiplier * len(results_of_second[8][dev][d])) + t), column=column).value = \
                        results_of_second[8][dev][d][t]
            column += 1
            if dev in ["PV", "CHP", ]:
                for method in ("use", "sell"):
                    rows = counter
                    ws2.cell(row=counter, column=column).value = dev + "_" + method
                    d_multiplier = -1
                    for d in range(len(results_of_second[8][dev, method])):
                        d_multiplier += 1
                        for t in range(len(results_of_second[8][dev, method][d])):
                            if d == 0 and t == 0:
                                rows += 1
                            ws2.cell(row=(rows + (d_multiplier * len(results_of_second[8][dev, method][d])) + t),
                                     column=column).value = \
                                results_of_second[8][dev, method][d][t]
                    column += 1

    for i in range(1, (ws2.max_row + 1)):  # ws2 has fewer rows than ws1 so it will work
        ws2.row_dimensions[i].height = 15
    for i in ("A", "B", "F", "G", "K", "L", "M", "N", "O", "P", "Q"):
        ws2.column_dimensions[i].width = 15
    for i in ("C", "D", "E", "H", "I", "J"):
        ws2.column_dimensions[i].width = 25

    name_of_output = name_of_output
    current_date = datetime.datetime.now()
    wb2.save(str(current_date.month) + "." + str(current_date.day) + "." + str(current_date.year) + " - " +
             str(current_date.hour) + "_" + str(current_date.minute) + "_" + "_" + name_of_output + ".xlsx")

