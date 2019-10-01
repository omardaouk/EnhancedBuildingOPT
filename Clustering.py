# Stuff to test

import numpy as np
import math
import gurobipy as gp
from openpyxl import Workbook
from openpyxl.chart import (AreaChart, ScatterChart, Reference, Series,)
from openpyxl.styles import (Font, Border, Side, Alignment)
import datetime



# Implementation of the k-medoids problem, as it is applied in
# Selection of typical demand days for CHP optimization
# Fernando Domínguez-Muñoz, José M. Cejudo-López, Antonio Carrillo-Andrés and
# Manuel Gallardo-Salazar
# Energy and Buildings. Vol 43, Issue 11 (November 2011), pp. 3036-3043

# Original formulation (hereafter referred to as [1]) can be found in:

# Integer Programming and the Theory of Grouping
# Hrishikesh D. Vinod
# Journal of the American Statistical Association. Vol. 64, No. 326 (June 1969)
# pp. 506-519
# Stable URL: http://www.jstor.org/stable/2283635

def k_medoids(distances, number_clusters, timelimit=100, mipgap=0.0001):
    """
    Parameters
    ----------
    distances : 2d array
        Distances between each pair of node points. `distances` is a
        symmetrical matrix (dissimmilarity matrix).
    number_clusters : integer
        Given number of clusters.
    timelimit : integer
        Maximum time limit for the optimization.
    """

    # Distances is a symmetrical matrix, extract its length
    length = distances.shape[0]

    # Create model
    model = gp.Model("k-Medoids-Problem")

    # Create variables
    x = {}  # Binary variables that are 1 if node i is assigned to cluster j
    y = {}  # Binary variables that are 1 if node j is chosen as a cluster
    for j in range(length):
        y[j] = model.addVar(vtype="B", name="y_" + str(j))

        for i in range(length):
            x[i, j] = model.addVar(vtype="B", name="x_" + str(i) + "_" + str(j))

    # Update to introduce the variables to the model
    model.update()

    # Set objective - equation 2.1, page 509, [1]
    obj = gp.quicksum(distances[i, j] * x[i, j]
                      for i in range(length)
                      for j in range(length))
    model.setObjective(obj, gp.GRB.MINIMIZE)

    # s.t.
    # Assign all nodes to clusters - equation 2.2, page 509, [1]
    # => x_i cannot be put in more than one group at the same time
    for i in range(length):
        model.addConstr(sum(x[i, j] for j in range(length)) == 1)

    # Maximum number of clusters - equation 2.3, page 509, [1]
    model.addConstr(sum(y[j] for j in range(length)) == number_clusters)

    # Prevent assigning without opening a cluster - equation 2.4, page 509, [1]
    for i in range(length):
        for j in range(length):
            model.addConstr(x[i, j] <= y[j])

    for j in range(length):
        model.addConstr(x[j, j] >= y[j])

    # Sum of main diagonal has to be equal to the number of clusters:
    model.addConstr(sum(x[j, j] for j in range(length)) == number_clusters)

    # Set solver parameters
    model.Params.TimeLimit = timelimit
    model.Params.MIPGap = mipgap

    # Solve the model
    model.optimize()

    # Get results
    r_x = np.array([[x[i, j].X for j in range(length)]
                    for i in range(length)])

    r_y = np.array([y[j].X for j in range(length)])

    r_obj = model.ObjVal

    return (r_y, r_x.T, r_obj)









def _distances(values, norm=2):
    """
    Compute distance matrix for all data sets (rows of values)

    Parameters
    ----------
    values : 2-dimensional array
        Rows represent days and columns values
    norm : integer, optional
        Compute the distance according to this norm. 2 is the standard
        Euklidean-norm.

    Return
    ------
    d : 2-dimensional array
        Distances between each data set
    """
    # Initialize distance matrix
    d = np.zeros((values.shape[1], values.shape[1]))

    # Define a function that computes the distance between two days
    dist = (lambda day1, day2, r:
            math.pow(np.sum(np.power(np.abs(day1 - day2), r)), 1 / r))

    # Remember: The d matrix is symmetrical!
    for i in range(values.shape[1]):  # loop over first days
        for j in range(i + 1, values.shape[1]):  # loop second days
            d[i, j] = dist(values[:, i], values[:, j], norm)

    # Fill the remaining entries
    d = d + d.T

    return d


def cluster(inputs, number_clusters=12, norm=2, time_limit=300, mip_gap=0.0,
            weights=None):
    """
    Cluster a set of inputs into clusters by solving a k-medoid problem.

    Parameters
    ----------
    inputs : 2-dimensional array
        First dimension: Number of different input types.
        Second dimension: Values for each time step of interes.
    number_clusters : integer, optional
        How many clusters shall be computed?
    norm : integer, optional
        Compute the distance according to this norm. 2 is the standard
        Euklidean-norm.
    time_limit : integer, optional
        Time limit for the optimization in seconds
    mip_gap : float, optional
        Optimality tolerance (0: proven global optimum)
    weights : 1-dimensional array, optional
        Weight for each input. If not provided, all inputs are treated equally.

    Returns
    -------
    scaled_typ_days :
        Scaled typical demand days. The scaling is based on the annual demands.
    nc : array_like
        Weighting factors of each cluster
    z : 2-dimensional array
        Mapping of each day to the clusters
    """
    # Determine time steps per day
    len_day = int(inputs.shape[1] / 365)

    # Set weights if not already given
    if weights == None:
        weights = np.ones(inputs.shape[0])
    elif not sum(weights) == 1:  # Rescale weights
        weights = np.array(weights) / sum(weights)

    # Manipulate inputs
    # Initialize arrays
    inputsTransformed = []
    inputsScaled = []
    inputsScaledTransformed = []

    # Fill and reshape
    # Scaling to values between 0 and 1, thus all inputs shall have the same
    # weight and will be clustered equally in terms of quality
    for i in range(inputs.shape[0]):
        vals = inputs[i, :]
        temp = ((vals - np.min(vals)) / (np.max(vals) - np.min(vals))
                * math.sqrt(weights[i]))
        inputsScaled.append(temp)
        inputsScaledTransformed.append(temp.reshape((len_day, 365), order="F"))
        inputsTransformed.append(vals.reshape((len_day, 365), order="F"))

    # Put the scaled and reshaped inputs together
    L = np.concatenate(tuple(inputsScaledTransformed))

    # Compute distances
    d = _distances(L, norm)

    # Execute optimization model
    (y, z, obj) = k_medoids(d, number_clusters, time_limit, mip_gap)

    # Section 2.3 and retain typical days
    nc = np.zeros_like(y)
    typicalDays = []

    # nc contains how many days are there in each cluster
    nc = []
    for i in range(len(y)):
        temp = np.sum(z[i, :])
        if temp > 0:
            nc.append(temp)
            typicalDays.append([ins[:, i] for ins in inputsTransformed])

    typicalDays = np.array(typicalDays)
    nc = np.array(nc, dtype="int")
    nc_cumsum = np.cumsum(nc) * len_day

    # Construct (yearly) load curves
    # ub = upper bound, lb = lower bound
    clustered = np.zeros_like(inputs)
    for i in range(len(nc)):
        if i == 0:
            lb = 0
        else:
            lb = nc_cumsum[i - 1]
        ub = nc_cumsum[i]

        for j in range(len(inputsTransformed)):
            clustered[j, lb:ub] = np.tile(typicalDays[i][j], nc[i])

    # Scaling to preserve original demands
    sums_inputs = [np.sum(inputs[j, :]) for j in range(inputs.shape[0])]
    scaled = np.array([nc[day] * typicalDays[day, :, :]
                       for day in range(number_clusters)])
    sums_scaled = [np.sum(scaled[:, j, :]) for j in range(inputs.shape[0])]
    scaling_factors = [sums_inputs[j] / sums_scaled[j]
                       for j in range(inputs.shape[0])]
    scaled_typ_days = [scaling_factors[j] * typicalDays[:, j, :]
                       for j in range(inputs.shape[0])]

    return (scaled_typ_days, nc, z,obj)


def cluster_weeks(inputs, number_clusters=12, norm=2, time_limit=300, mip_gap=0.0,
            weights=None):
    """
    Cluster a set of inputs into clusters by solving a k-medoid problem.

    Parameters
    ----------
    inputs : 2-dimensional array
        First dimension: Number of different input types.
        Second dimension: Values for each time step of interes.
    number_clusters : integer, optional
        How many clusters shall be computed?
    norm : integer, optional
        Compute the distance according to this norm. 2 is the standard
        Euklidean-norm.
    time_limit : integer, optional
        Time limit for the optimization in seconds
    mip_gap : float, optional
        Optimality tolerance (0: proven global optimum)
    weights : 1-dimensional array, optional
        Weight for each input. If not provided, all inputs are treated equally.

    Returns
    -------
    scaled_typ_days :
        Scaled typical demand days. The scaling is based on the annual demands.
    nc : array_like
        Weighting factors of each cluster
    z : 2-dimensional array
        Mapping of each day to the clusters
    """
    # Determine time steps per day
    len_week = int(inputs.shape[1] / 52)

    # Set weights if not already given
    if weights == None:
        weights = np.ones(inputs.shape[0])
    elif not sum(weights) == 1:  # Rescale weights
        weights = np.array(weights) / sum(weights)

    # Manipulate inputs
    # Initialize arrays
    inputsTransformed = []
    inputsScaled = []
    inputsScaledTransformed = []

    # Fill and reshape
    # Scaling to values between 0 and 1, thus all inputs shall have the same
    # weight and will be clustered equally in terms of quality
    for i in range(inputs.shape[0]):
        vals = inputs[i, :]
        temp = ((vals - np.min(vals)) / (np.max(vals) - np.min(vals))
                * math.sqrt(weights[i]))
        inputsScaled.append(temp)
        inputsScaledTransformed.append(temp.reshape((len_week, 52), order="F"))
        inputsTransformed.append(vals.reshape((len_week, 52), order="F"))

    # Put the scaled and reshaped inputs together
    L = np.concatenate(tuple(inputsScaledTransformed))

    # Compute distances
    d = _distances(L, norm)

    # Execute optimization model
    (y, z, obj) = k_medoids(d, number_clusters, time_limit, mip_gap)

    # Section 2.3 and retain typical days
    nc = np.zeros_like(y)
    typicalDays = []

    # nc contains how many days are there in each cluster
    nc = []
    for i in range(len(y)):
        temp = np.sum(z[i, :])
        if temp > 0:
            nc.append(temp)
            typicalDays.append([ins[:, i] for ins in inputsTransformed])

    typicalDays = np.array(typicalDays)
    nc = np.array(nc, dtype="int")
    nc_cumsum = np.cumsum(nc) * len_week

    # Construct (yearly) load curves
    # ub = upper bound, lb = lower bound
    clustered = np.zeros_like(inputs)
    for i in range(len(nc)):
        if i == 0:
            lb = 0
        else:
            lb = nc_cumsum[i - 1]
        ub = nc_cumsum[i]

        for j in range(len(inputsTransformed)):
            clustered[j, lb:ub] = np.tile(typicalDays[i][j], nc[i])

    # Scaling to preserve original demands
    sums_inputs = [np.sum(inputs[j, :]) for j in range(inputs.shape[0])]
    scaled = np.array([nc[day] * typicalDays[day, :, :]
                       for day in range(number_clusters)])
    sums_scaled = [np.sum(scaled[:, j, :]) for j in range(inputs.shape[0])]
    scaling_factors = [sums_inputs[j] / sums_scaled[j]
                       for j in range(inputs.shape[0])]
    scaled_typ_weeks = [scaling_factors[j] * typicalDays[:, j, :]
                       for j in range(inputs.shape[0])]

    return (scaled_typ_weeks, nc, z, obj)







def calculate_clusters(inputs_clustering,number_clusters,norm=2,mip_gap=0.0,weights = None):
    (inputs, nc, z, obj) = cluster(inputs_clustering,
                                   number_clusters=number_clusters,
                                   norm=2,
                                   mip_gap=0.0,
                                   weights=weights)
    return (inputs, nc, z,obj)

def calculate_clusters_weeks(inputs_clustering,number_clusters,norm=2,mip_gap=0.0,weights = None):
    (inputs, nc, z, obj) = cluster_weeks(inputs_clustering,
                                   number_clusters=number_clusters,
                                   norm=2,
                                   mip_gap=0.0,
                                   weights=weights)
    return (inputs, nc, z,obj)


def error_vs_number_of_clusters(inputs_clustering , number_typtage):
    '''The inputs should be in the form of an array of lists of 8760 values ... like
    inputs_clustering = np.array([raw_inputs["electricity"],
                              #raw_inputs["dhw"],
                              #raw_inputs["sh"],
                              #raw_inputs["solar_irrad"],
                              #raw_inputs["temperature"]
                              ])'''

    objective_values = {}
    maximums = {}
    minimums = {}

    num_typ_Tage = number_typtage

    for i in range(num_typ_Tage):
        print("################")
        print("################")
        print("Step", str(i + 1), "of", str(num_typ_Tage))
        print("################")
        print("################")
        number_clusters = i + 1
        #(inputs, nc, z, obj) = cluster(inputs_clustering,
        #                               number_clusters=number_clusters,
        #                               norm=2,
        #                               mip_gap=0.0,
        #                               weights=[1, 1, 1, 1, 0])

        (inputs, nc, z, obj) = cluster_weeks(inputs_clustering,
                                       number_clusters=number_clusters,
                                       norm=2,
                                       mip_gap=0.0,
                                       weights=[1, 1, 1, 1/2, 1/2, 1/2, 1, 1, 1/2, 1])#[1,1,1,1])


        objective_values[i + 1] = obj
        maximums[i + 1] = max(nc)
        minimums[i + 1] = min(nc)

    wb2 = Workbook()
    ws2 = wb2.create_sheet('Clustering', 0)
    ws2.cell(row=1, column=1).value = 'Number of Data Groups to Cluster'
    ws2.cell(row=1, column=2).value = 'Number of Clusters'
    ws2.cell(row=1, column=3).value = 'Smallest Cluster'
    ws2.cell(row=1, column=4).value = 'Largest Cluster'
    ws2.cell(row=1, column=5).value = 'Value of the Objective Function'
    ws2.cell(row=2, column=1).value = inputs_clustering.shape[0]

    for i in range(num_typ_Tage):
        ws2.cell(row=i + 2, column=2).value = i + 1
        ws2.cell(row=i + 2, column=3).value = minimums[i + 1]
        ws2.cell(row=i + 2, column=4).value = maximums[i + 1]
        ws2.cell(row=i + 2, column=5).value = objective_values[i + 1]

    chart1 = ScatterChart()
    chart1.title = "Error vs. #Clusters"
    chart1.style = 13
    chart1.x_axis.title = 'Number of Clusters'
    chart1.y_axis.title = 'Value of the Objective Function(Distances to be minimized)'
    chart1.width = 15
    chart1.height = 15

    xvalues = Reference(ws2, min_col=2, min_row=2, max_row=num_typ_Tage + 1)
    values = Reference(ws2, min_col=5, min_row=1, max_row=num_typ_Tage + 1)
    series = Series(values, xvalues, title_from_data=True)
    chart1.series.append(series)
    chart1.legend.position = 'b'

    ws2.add_chart(chart1, "G2")

    chart2 = ScatterChart()
    chart2.title = "Cluster Sizes"
    chart2.style = 13
    chart2.x_axis.title = 'Number of Clusters'
    chart2.y_axis.title = 'Cluster Size'
    chart2.width = 15
    chart2.height = 15
    chart2.legend.position = 'b'

    xvalues = Reference(ws2, min_col=2, min_row=2, max_row=num_typ_Tage + 1)
    for i in (3, 4):
        values = Reference(ws2, min_col=i, min_row=1, max_row=num_typ_Tage + 1)
        series = Series(values, xvalues, title_from_data=True)
        chart2.series.append(series)

    ws2.add_chart(chart2, "Q2")

    for i in range(1, (ws2.max_row + 1)):  # ws2 has fewer rows than ws1 so it will work
        ws2.row_dimensions[i].height = 15
    for i in ('A', 'B', 'C', 'D', 'E'):
        ws2.column_dimensions[i].width = 30

    name_of_output = "Error vs Number of Clusters (clustering CO2(t))" + str(num_typ_Tage)
    current_date = datetime.datetime.now()
    wb2.save(str(current_date.month) + "." + str(current_date.day) + "." + str(current_date.year) + " - " +
             str(current_date.hour) + "_" + str(current_date.minute) + "_" + "_" + name_of_output + ".xlsx")



if __name__ == "__main__":


    raw_inputs = {}

    # # Scale inputs to kW
    # """"HERE I NEED TO WRITE THE NAMES OF MY INPUT FILES INSTEAD ... CO2(t), Electricity demand ... etc.
    # In his case house was something in the loop ... for house in bla  bla  bla ... so it worked ... here i added
    # "" and made it a string so that it does not give an error ... but i will remove it in anycase so it is not a
    # problem"""
    #raw_inputs["CO2(t)"] = np.loadtxt("CO2(t)_2018_8760.txt")
    #raw_inputs["dhw"] = np.loadtxt("raw_inputs/building_" + "house" + "/dhw.csv")
    #raw_inputs["sh"] = np.loadtxt("raw_inputs/building_" + "house" + "/space_heating.csv")
    #raw_inputs["solar_irrad"] = np.loadtxt("raw_inputs/building_" + "house" + "/solar_rad_35deg.csv") / 1000
    #raw_inputs["solar_irrad"] = np.maximum(raw_inputs["solar_irrad"], 0)
    #raw_inputs["temperature"] = np.loadtxt("raw_inputs/building_" + "house" + "/temperature.csv"

    #inputs_clustering = np.array([raw_inputs["CO2(t)"],
    #                              #raw_inputs["dhw"],
    #                              #raw_inputs["sh"],
    #                              #raw_inputs["solar_irrad"],
    #                              #raw_inputs["temperature"]
    #                              ])
    """ For Typ Tage"""
    # raw_inputs["CO2(t)"] = np.loadtxt("CO2(t)_2018_8760.txt")
    # raw_inputs["dhw"] = np.loadtxt("input_data/demand_domestic_hot_water" + ".txt")
    # raw_inputs["electricity"] = np.loadtxt("electricity_sfh_4_medium.txt")
    # raw_inputs["chp_rem"] = np.loadtxt("CHP_Remuneration.txt")
    # #raw_inputs["sun_rad_0"] = SunRad[0]
    # #raw_inputs["sun_rad_1"] = SunRad[1]
    # #raw_inputs["sun_rad_2"] = SunRad[2]
    # #raw_inputs["sun_rad_3"] = SunRad[3]
    # #raw_inputs["sun_rad_4"] = SunRad[4]
    # #raw_inputs["sun_rad_5"] = SunRad[5]
    # #raw_inputs["T_e_raw"] = T_e_raw
    #
    # inputs_clustering = np.array([raw_inputs["CO2(t)"],
    #                               raw_inputs["dhw"],
    #                               raw_inputs["electricity"],
    #                               raw_inputs["chp_rem"],
    #                               # raw_inputs["sun_rad"], Problem here because Sun Rad is not a list of 8760 values
    #                               # but rather an array of 6 lists each containing 8760 values ...
    #                               # so it needs to be split into 6 lists that either all go into the clustering as
    #                               # separate cluster input ... or just the ones that are later used in the code
    #                               # for now i will put them all as individual lists and later on i can filter out
    #                               # some of them if i realize that they are not really used for anything
    #                               #raw_inputs["sun_rad_0"],
    #                               #raw_inputs["sun_rad_1"],
    #                               #raw_inputs["sun_rad_2"],
    #                               #raw_inputs["sun_rad_3"],
    #                               #raw_inputs["sun_rad_4"],
    #                               #raw_inputs["sun_rad_5"],
    #                               #raw_inputs["T_e_raw"],
    #                               # raw_inputs["sh"],
    #                               # raw_inputs["solar_irrad"],
    #                               # raw_inputs["temperature"]
    #                               ])

    """ For Typ Wochen"""
    raw_inputs["CO2(t)"] = np.loadtxt("CO2(t)_2018_weeks.txt")
    raw_inputs["dhw"] = np.loadtxt("input_data/demand_domestic_hot_water 8736" + ".txt")
    raw_inputs["electricity"] = np.loadtxt("electricity_sfh_4_medium 8736.txt")
    raw_inputs["chp_rem"] = np.loadtxt("CHP_Remuneration 8736.txt ")


    inputs_clustering = np.array([raw_inputs["CO2(t)"],
                                  #raw_inputs["dhw"],
                                  #raw_inputs["electricity"],
                                  #raw_inputs["chp_rem"],
                                  # raw_inputs["sun_rad"], Problem here because Sun Rad is not a list of 8760 values
                                  # but rather an array of 6 lists each containing 8760 values ...
                                  # so it needs to be split into 6 lists that either all go into the clustering as
                                  # separate cluster input ... or just the ones that are later used in the code
                                  # for now i will put them all as individual lists and later on i can filter out
                                  # some of them if i realize that they are not really used for anything
                                  # raw_inputs["sun_rad_0"],
                                  # raw_inputs["sun_rad_1"],
                                  # raw_inputs["sun_rad_2"],
                                  # raw_inputs["sun_rad_3"],
                                  # raw_inputs["sun_rad_4"],
                                  # raw_inputs["sun_rad_5"],
                                  # raw_inputs["T_e_raw"],
                                  # raw_inputs["sh"],
                                  # raw_inputs["solar_irrad"],
                                  # raw_inputs["temperature"]
                                  ])

    number_typtage = 365
    number_typwochen = 52

    #error_vs_number_of_clusters(inputs_clustering, number_typtage)
    error_vs_number_of_clusters(inputs_clustering, number_typwochen)



# # Determine time steps per day
# len_day = int(inputs_clustering.shape[1] / 365)
#
# clustered = {}
# clustered["electricity"] = inputs[0]
# #clustered["dhw"] = inputs[1]
# #clustered["sh"] = inputs[2]
# #clustered["solar_irrad"] = inputs[3]
# #clustered["temperature"] = inputs[4]
# #clustered["design_heat_load"] = design_heat_load
# clustered["weights"] = nc
# clustered["z"] = z
#
# # print("The representative clusters are")
# # print(clustered["electricity"])
# # print("##################")
# # print("##################")
# # print(clustered["electricity"].shape)
# # print("##################")
# # print("##################")
# # print("The list containing the number of weight days within each cluster is")
# # print(clustered["weights"])
# # print(clustered["weights"].shape)
# # print("min is " + str(min(clustered["weights"])))
# # print("max is " + str(max(clustered["weights"])))
# # print("##################")
# # print("##################")
# # print("The objective value is")
# # print(obj)
# # print("##################")
# # print("##################")
# #print("The matrix showimg which days belong to which cluster is")
# #print(clustered["z"])
# #print(clustered["z"].shape)
# #print("##################")
# #print("##################")
#
# test_results = {}
# for i in range(clustered["z"].shape[0]):
#     list=[]
#     for j in range(clustered["z"].shape[1]):
#         if clustered["z"][i][j] == 1:
#             list.append((i,j))
#         else:
#             continue
#     test_results[i] = list
#
# for i in range (clustered["z"].shape[0]):
#     if len(test_results[i]) != 0:
#         # print("##################")
#         # print("##################")
#         # print(i)
#         # print(test_results[i])
#         # print("Contains " + str(len(test_results[i])) + " elements")
#         # print("Element " + str(i) + " represents the following " + str(len(test_results[i])) + " days :")
#         blabla=[]
#         for j in range(len(test_results[i])):
#             blabla.append(test_results[i][j][1])
#         # print(blabla)


