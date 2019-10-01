#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Tue May 26 10:50:23 2015

@author: tsz
"""

import gurobipy as gp
import numpy as np
import sun
import input_data.components as components
import input_data.building as building
import pickle
from openpyxl import load_workbook  # My Addition
from Clustering import calculate_clusters

wb1 = load_workbook('CO2(t)_2018_8760.xlsx', data_only=True)  # My Addition
ws1 = wb1['2018']  # My Addition

wb11 = load_workbook('CO2(t)_2030_8760.xlsx', data_only=True)  # My Addition
ws11 = wb11['2030']  # My Addition

def optimize(max_emissions, options):

    # Constants
    direction = ("south", "west", "north", "east", "roof", "floor")
    direction2 = ("wall", "roof", "floor")
    direction3 = ("south", "west", "north", "east")
    direction4 = ("roof", "floor")
    
    # Time relevant parameters
    timesteps = 24  # Time steps per typical day
    weight_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # Days per month
    number_days = len(weight_days)  # Number of typical days per year
    monthsCum0 = np.array([0] + weight_days[0:number_days-1])
    monthsCum0 = monthsCum0.cumsum()
    dt = 1  # Time step width in hours
    
    # 5R1C ISO 13790 Parameter
    # Constants for calculation of A_m, dependent of building class
    # (DIN EN ISO 13790, section 12.3.1.2, page 81, table 12)
    f_class = {}
    f_class["Am"] = [2.5, 2.5, 2.5, 3.0, 3.5]  # m²
    # specific heat transfer coefficient 
    # (DIN EN ISO 13790, section 7.2.2.2, page 35)
    h_is = 3.45  # [W/(m²K)]
    # non-dimensional relation between the area of all indoor surfaces and the
    # effective floor area A["f"]
    # (DIN EN ISO 13790, section 7.2.2.2, page 36)
    lambda_at = 4.5
    # specific heat transfer coefficient
    # (DIN EN ISO 13790, section 12.2.2, page 79)
    h_ms = 9.1   # [W/(m²K)]
    ###
    ###
    # building parameters
    # South, West, North, East, Roof, PV/STC
    beta = building.beta    # slope angle 
    gamma = building.gamma  # surface azimuth angle
    
    # Building geometry 
    A = building.A  # [m2]
    V = building.V  # [m3] (A_f*heightWalls)
    
    # Form factor for radiation between the element and the sky
    # (DIN EN ISO 13790, section 11.4.6, page 73)
    F_r = building.F_r
    
    # Internal gains in W
    phi_int = building.phi_int
    
    # Power generation / consumption in kW (positive, device dependent)
    power = building.power  
    
    # Temperature bounds and ventilation input
    T_set_min = building.T_set_min              # °C
    ventilationRate = building.ventilationRate  # [airchanges per hour]
    
    # Approximated room temperature in °C
    T_i_appr = building.T_i_appr
    ###
    ###
    # Data depending on location (Aachen)
    ####################################
    # Missing: T_ne, T_me, f_g1, G_w, cp_air, rho_air
    weatherData = np.loadtxt("input_data/weather.txt", skiprows=2, delimiter="\t")
    weatherData = weatherData[0:8760, :]
    location = (49.5, 8.5)
    altitude = 0  # m, section 5.2.1.6.1, page 16
    timeZone = 1
    SunDirect = weatherData[:, 14]
    SunDiffuse = weatherData[:, 15]
    T_e_raw = weatherData[:, 9]  # Dry-Bulb Temperature
    albedo = 0.2  # Ground reflectance
    # Mean difference between outdoor temperature and the apparent sky-temperature
    # (DIN EN ISO 13790, section 11.4.6,  page 73)
    Delta_theta_er = 11  # [K]
    
    # Computation of design heat load 
    T_ne = -23.33  # [°C] outside design temperature
    T_me = 10.08  # [°C] outside average temperature
    f_g1 = 1.45
    # Reduction factor
    f_g2 = (T_set_min - T_me) / (T_set_min - T_ne)
    G_w = 1.0  # influence of ground water neglected
    c_p_air = 1000.0  # [J/kgK]
    rho_air = 0.986  # [kg/m3]
    
    # Outdoor conditions (Sun radiation on surfaces and temperature)
    SunRad = sun.getSolarGains(0, 3600, 8760,
                               timeZone=timeZone,
                               location=location,
                               altitude=altitude,
                               beta=beta,
                               gamma=gamma,
                               beam=SunDirect,
                               diffuse=SunDiffuse,
                               albedo=albedo)
    print("the SunRad is an output we get from using the weather ... and it looks like this ")
    print(SunRad)
    print(SunRad.shape)
    print("The dry bulb temperature is also an output we get from using the weather ... and it looks like this ")
    print(T_e_raw)
    print(T_e_raw.shape)

    raw_inputs = {}

    raw_inputs["CO2(t)"] = np.loadtxt("CO2(t)_2018_8760.txt")
    raw_inputs["dhw"] = np.loadtxt("input_data/demand_domestic_hot_water" + ".txt")
    raw_inputs["electricity"] = np.loadtxt("electricity_sfh_4_medium.txt")
    raw_inputs["chp_rem"] = np.loadtxt("CHP_Remuneration.txt")
    raw_inputs["sun_rad_0"] = SunRad[0]
    raw_inputs["sun_rad_1"] = SunRad[1]
    raw_inputs["sun_rad_2"] = SunRad[2]
    raw_inputs["sun_rad_3"] = SunRad[3]
    raw_inputs["sun_rad_4"] = SunRad[4]
    raw_inputs["sun_rad_5"] = SunRad[5]
    raw_inputs["T_e_raw"] = T_e_raw
    # raw_inputs["sh"] = np.loadtxt("raw_inputs/building_" + "house" + "/space_heating.csv")
    # raw_inputs["solar_irrad"] = np.loadtxt("raw_inputs/building_" + "house" + "/solar_rad_35deg.csv") / 1000
    # raw_inputs["solar_irrad"] = np.maximum(raw_inputs["solar_irrad"], 0)
    # raw_inputs["temperature"] = np.loadtxt("raw_inputs/building_" + "house" + "/temperature.csv"

    inputs_clustering = np.array([raw_inputs["CO2(t)"],
                                  raw_inputs["dhw"],
                                  raw_inputs["electricity"],
                                  raw_inputs["chp_rem"],
                                  #raw_inputs["sun_rad"], Problem here because Sun Rad is not a list of 8760 values
                                  # but rather an array of 6 lists each containing 8760 values ...
                                  # so it needs to be split into 6 lists that either all go into the clustering as
                                  # separate cluster input ... or just the ones that are later used in the code
                                  # for now i will put them all as individual lists and later on i can filter out
                                  # some of them if i realize that they are not really used for anything
                                  raw_inputs["sun_rad_0"],
                                  raw_inputs["sun_rad_1"],
                                  raw_inputs["sun_rad_2"],
                                  raw_inputs["sun_rad_3"],
                                  raw_inputs["sun_rad_4"],
                                  raw_inputs["sun_rad_5"],
                                  raw_inputs["T_e_raw"],
                                  # raw_inputs["sh"],
                                  # raw_inputs["solar_irrad"],
                                  # raw_inputs["temperature"]
                                  ])

    print("##############")
    print("The 'inputs_clustering' that we are sending to the cluster function are of the form :")
    print(inputs_clustering)
    print("The 'Inputs' has the form :  " + str(inputs_clustering.shape))
    print("##############")

    for i in range(6):
        print("##############")
        print("Sun Rad " + str(i) + " is of the the form :")
        print(raw_inputs["sun_rad_" + str(i)])
        print("Sun Rad " + str(i) + "has the form :  " + str(raw_inputs["sun_rad_" + str(i)].shape))
        print("##############")

    print('For raw_inputs["CO2(t)"] : ')
    print('The type is :  ' + str(type(raw_inputs["CO2(t)"])))
    print('The shape is :  ' + str(raw_inputs["CO2(t)"].shape))
    print('The string form is :  ' + str(raw_inputs["CO2(t)"]))
    print("##############")
    print('For raw_inputs["sun_rad_5"] : ')
    print('The type is :  ' + str(type(raw_inputs["sun_rad_5"])))
    print('The shape is :  ' + str(raw_inputs["sun_rad_5"].shape))
    print('The string form is :  ' + str(raw_inputs["sun_rad_5"]))

    number_clusters = 2

    print("##############")
    print("##############")
    print("Bilding the Clusters")
    print("##############")
    print("##############")
    weights_of_input = [1,1,1,1,1,1,1,1,1,1,1]
    (inputs, nc, z, obj) = calculate_clusters(inputs_clustering,number_clusters,norm=2,mip_gap=0.0, weights = weights_of_input)

    clustered = {}
    clustered["representative_days"] = {}
    clustered["representative_days"]["CO2(t)"] = inputs[0]
    clustered["representative_days"]["dhw"] = inputs[1]
    clustered["representative_days"]["electricity"] = inputs[2]
    clustered["representative_days"]["chp_rem"] = inputs[3]
    clustered["representative_days"]["sun_rad_0"] = inputs[4]
    clustered["representative_days"]["sun_rad_1"] = inputs[5]
    clustered["representative_days"]["sun_rad_2"] = inputs[6]
    clustered["representative_days"]["sun_rad_3"] = inputs[7]
    clustered["representative_days"]["sun_rad_4"] = inputs[8]
    clustered["representative_days"]["sun_rad_5"] = inputs[9]
    clustered["representative_days"]["T_e_raw"] = inputs[10]
    #clustered["design_heat_load"] = design_heat_load
    clustered["weights"] = nc
    clustered["z"] = z
    clustered["weights_of_input"] = weights_of_input
    clustered["obj"] = obj


    print("##############")
    print("##############")
    print("Clustering Stuff Here")
    print("##############")
    print("##############")
    print("The 'Inputs' result that we get is :")
    print(inputs)
    print("The 'Inputs' has " + str(len(inputs)) + " elements .... one for each input data that we wanted to cluster")
    print("Each element has " + str(len(inputs[0])) + " sub_elements ... one for each cluster or Typ Tag")
    print("Each sub element has " + str(len(inputs[0][0])) + " values ... one for each hour of the day")
    print("##############")
    print("The 'Weights' of the result that we get is :")
    print(nc)
    print("The 'nc' has the form :  " + str(nc.shape))
    print("##############")


    """
    def calculate_clusters(inputs_clustering,number_clusters,norm=2,mip_gap=0.0,weights=[1, 1, 1, 1, 0]):
    (inputs, nc, z, obj) = cluster(inputs_clustering,
                                   number_clusters=number_clusters,
                                   norm=2,
                                   mip_gap=0.0,
                                   weights=[1, 1, 1, 1, 0])
    return (inputs, nc, z,obj)
    """
                                           
    T_e = np.zeros((number_days, timesteps))
    SunRad_part = np.zeros((6, number_days, timesteps))
    
    for d in range(number_days):
        for t in range(timesteps):
            for i in range(weight_days[d]):
                T_e[d, t] += T_e_raw[i*timesteps+t+monthsCum0[d]*timesteps] / weight_days[d]
                SunRad_part[:, d, t] += (SunRad[:, i*timesteps+t+monthsCum0[d]*timesteps] / weight_days[d] / 1000)  # kW
    ###

    print("the SunRad_part is what SunRad is used to calculate and it looks like this ... ")
    print(SunRad_part)
    print(SunRad_part.shape)
    ###
    # Building envelope components
    ##############################
    t_life = components.t_life  # Economic lifetime of different technologies [3]
    U = components.U  # thermal transmittance [kW/(m²*K)]
    epsilon = components.epsilon  # radiation emission coefficient [-]
    R_se = components.R_se  # surface resistance [m²*K/kW]
    alpha_Sc = components.alpha_Sc  # adsorption coefficient [-]
    kappa = components.kappa  # specific heat storage capacity [kWh/(K*m²]
    g_gl = components.g_gl  # overall energy transmittance of glazings
    inv = components.inv  # specific investion costs in [€/m²]
    
    # ratio of window-frame
    # (DIN EN ISO 13790, section 11.4.5, page 73)
    F_F = 0 
    
    # thermal radiation transfer
    # [kW/(m²*K)] DIN EN ISO 13790, section 11.4.6, page 73
    h_r_factor = 5.0  # W / (m²K)
    h_r = {}
    h_r["opaque", "wall"] = h_r_factor*np.array(epsilon["opaque", "wall"])
    h_r["opaque", "roof"] = h_r_factor*np.array(epsilon["opaque", "roof"])
    h_r["opaque", "floor"] = h_r_factor*np.array(epsilon["opaque", "floor"])
    h_r["window"] = h_r_factor*np.array(epsilon["window"])
    
    # thermal transmittance coefficient H_ve [W/K]
    # (DIN EN ISO 13790, section 9.3.1, equation 21, page 49)
    H_ve = rho_air * c_p_air * ventilationRate * V / 3600
    
    # thermal transmittance coefficient H_tr_is [W/K]
    # (DIN EN ISO 13790, section 7.2.2.2, equation 9, page 35)
    A_tot = lambda_at * A["f"]
    H_tr_is = h_is * A_tot
    
    # heat flow phi_ia [W] 
    # (DIN EN ISO 13790, section C2, page 110, eq. C.1)
    phi_ia = 0.5 * phi_int
    
    # matching coefficient for thermal transmittance coefficient if temperature is 
    # unequal to T_e, otherwise = 1
    # Assumption: Constant annual heat flow through ground
    # (ISO 13370 A.5 p 25 eq. A8)
    T_e_mon = np.mean(T_e, axis=1)  # monthly mean outside temperature
    b_floor = []       # ground p.44 ISO 13790
    
    T_i_year = 22.917  # annual mean indoor temperature
    T_e_year = 9.71  # annual mean outside temperature
    
    # Heating period from October until May (important for T_i_appr)
    for i in range(4):
        b_floor.append((T_i_year - T_e_year) / (T_i_appr[i] - T_e_mon[i]))
    for i in range(5):
        b_floor.append((T_i_year - T_e_year) / (T_i_appr[4+i] - T_e_mon[4+i]))
    for i in range(3):
        b_floor.append((T_i_year - T_e_year) / (T_i_appr[9+i] - T_e_mon[9+i]))
    b_floor = np.array(b_floor)
    
    b_tr = {}
    b_tr["wall"] = np.ones((number_days, timesteps))
    b_tr["roof"] = np.ones((number_days, timesteps))
    b_tr["floor"] = np.zeros((number_days, timesteps))
    for i in range(number_days):
        b_tr["floor"][i, :] = b_floor[i]
    
    # upper and lower bound for T_s for linearization
    T_s_o = 50  # [°C]
    T_s_u = 0   # [°C]
    
    # upper and lower bound for T_m for linearization
    T_m_o = 50  # [°C]
    T_m_u = 0   # [°C]
    
    # initial temperature for building envelope
    T_init = 20  # [°C]
    
    # shadow coefficient for sunblinds 
    #       (DIN EN ISO 13790, section 11.4.3, page 71)
    F_sh_gl = 1  # Assumption : no sunblinds (modelled manually, see below)
      
    # Dictionary for Irradiation to imitate sunblinds manually [kW/m²]
    I_sol = {}
    directions = ("south", "west", "north", "east", "roof")
    for drct3 in range(len(directions)):
        I_sol[directions[drct3]] = SunRad_part[drct3, :, :]
        I_sol["window", directions[drct3]] = SunRad_part[drct3, :, :]
    
    I_sol["floor"] = np.zeros_like(I_sol["roof"])
    I_sol["window", "floor"] = np.zeros_like(I_sol["roof"])
    
    limit_shut_blinds = 0.1  # kW/m²
    for d in range(4, 9):  # May until September
        for t in range(timesteps):
            for drct3 in range(len(directions)):  # For all directions
                if SunRad_part[drct3, d, t] > limit_shut_blinds:
                    I_sol["window", directions[drct3]][d, t] = 0.15 * SunRad_part[drct3, d, t]
    
    # 5R1C ISO 13790 End
    ###################
    
    # Economics
    ###########
    
    t_clc = 10.0  # Observation time period in years
    rate = 0.055  # Interest rate for annuity factor
    tax = 1.19   # Value added tax
    prChange = {"el": 1.057,  # Price change factors per year for electricity
                "gas": 1.041,  # Price change factors per year for natural gas
                "eex": 1.0252,  # Price change factors per year for eex compensation
                "infl": 1.017}  # Price change factors per year for inflation
    
    # Irradiation on 35° tilted surface (for "STC"-efficiency) in kW/m2
    solar_rad = SunRad_part[5, :, :]
    
    q = 1 + rate  # Abbreviation
    crf = (q ** t_clc * rate) / (q ** t_clc - 1)  # Capital recovery factor
    
    b = {key: (1 - (prChange[key] / q) ** t_clc) / (q - prChange[key])
         for key in prChange.keys()}  # Calculation of cash value B concerning energy price increase
    
    price = {"CHP": 0.0541,  # Subsidy for electricity produced by CHP units from German co-generation act in EUR/kWh
             "el": 0.2660,  # Electricity price in EUR/kWh
             ("CHP", "sell"): np.loadtxt("input_data/remuneration_chp.txt"),  # Remuneration for sold electricity from CHP in EUR/kWh
             ("PV",  "sell"): 0.1427,  # Remuneration for sold electricity from PV, fixed for 20 years
             ("gas", "CHP"): 0.0608,  # Natural gas price for CHP units without energy tax based on LHV in EUR/kWh
             ("gas", "Boiler"): 0.0693}  # Natural gas price for other applications based on LHV in EUR/kWh


    
    c_meter = {"gas": 157}  # Fixed costs for gas metering system EUR/year
    
    # Emission coefficients in kg_CO2 / kWh
    emi = {"PV":  0,
           "gas": 0.25,
           "el":  0.569}
    
    emission_max = max_emissions  # tons of CO2 per year

    # House related data:
    heat = {}   # Heat generation / consumption in kW (positive, device dependent)
    
    # Heat demand and domestic hot water demand in kW
    heat_dhw_raw = np.loadtxt("input_data/demand_domestic_hot_water.txt") / 1000
    
    # Monthly averaging
    heat["DHW"] = np.zeros_like(T_e)
    for d in range(number_days):
        for t in range(timesteps):
            heat["DHW"][d, t] = np.mean([heat_dhw_raw[t+timesteps*monthsCum0[d]+timesteps*i]
                                        for i in range(weight_days[d])])
    
    t_amb = T_e  # Ambient temperature for this region in °C
    
    # Common dictionaries for devices
    bounds = {}
    c_fix = {}
    eta = {}
    part_load = {}
    cap = {}
    f_serv = {} 
    
    # Parameters for each device
    # Battery

    # My Addition
    # Kontinuierliche Variablen

    inv["Battery"] = 592.1  # Specific investment costs Battery in EUR/kWh
    c_fix["Battery"] = 5137  # Fix cost share Battery in EUR
    t_life["Battery"] = 13.7  # Based on number of cycles]
    bounds["Battery", "up"] = 30  # dummy in kWh # 5000 if testing seasonal storage
    bounds["Battery", "charge"] = 0.6286 # 1 # 0.6286  # Specific charging and discharging rates in kW/kWh ... 0.6286 so that we can charge and discharge fully within 1.6 hours ... if we want within 1 hour simply make the rates 1
    bounds["Battery", "discharge"] = 0.6286 # 1 # 0.6286  # Specific charging and discharging rates in kW/kWh ... 0.6286 so that we can charge and discharge fully within 1.6 hours ... if we want within 1 hour simply make the rates 1
    bounds["Battery", "DOD"] = 0.8
    eta["Battery", "charge"] = 0.8875  # eta_oneWay = sqrt(eta_cycle)
    eta["Battery", "discharge"] = 0.8875  # eta_oneWay = sqrt(eta_cycle)
    sub_battery = {("PV", "sell"): 0.6,  # Sold PV electricity must be <= 0.6 * ratedPower
                   ("PV", "sub"): 0.3,  # Maximum subsidy of 30 % of the investments
                   ("PV", "investment"): 1600.0,  # Refund 1600 EUR / kW_peak PV
                   ("PV", "bound"): 2000.0}  # Maximum refund 2000 EUR / kW_peak PV
    init = {}  # The initial state of charge is defined at a latter point for the  battery


    # Thermal energy storage
    # After continuous TES
    # The "i=0" term is for the already existing device
    # The "i=1" term is for the continuous devices
    # inv["TES"] = [0, 5000]  # TO TEST TES COST EFFECT EUR and EUR/m3
    # inv["TES"] = [0, 1000000]  # TO TEST TES COST EFFECT EUR and EUR/m3
    inv["TES"] = [0, 726.9]  # EUR and EUR/m3
    cap["TES"] = [0.1, 2.0]  # capacity and maximum capacity in m³
    c_fix["TES"] = 447  # Fix cost share TES in EUR
    t_life["TES"] = 20
    bounds["TES", "up"] = max(cap["TES"])  # Upper bound for storage capacity in m3
    # bounds["TES", "lo"] = min(cap["TES"])  # Lower bound for storage capacity in m3
    bounds["TES", "lo"] = 0  # made it 0 because it was messing up my new runs because i want to
    # look at TES from 0 to max and because of this ... a restriction later is causing contradiction because it forces
    # The cap to be greater than 0
    deltaT = {"TES": 35,  # Maximum difference between upper and lower part of the storage unit in K
              "HP": 10}  # Reduced difference for heat pumps
    cp_water = 4180.0 / 3600000  # Specific heat capacity of water kWh/kgK
    rho_water = 975  # Density of water in kg/m3 at 75°C
    sto_loss = 0.006  # Storage losses 0.6 % per hour
    init["TES"] = 0  # Initialize storage content with 0 kWh

    # Boiler
    # After continuous Boiler
    # The "i=0" term is for the already existing device
    # The "i=1" or more term is for the continuous devices

    inv["Boiler"] = [0, 51.56]  # Investment costs for boiler in EUR and EUR/m3
    cap["Boiler"] = [18, 37.5]  # Capacity and maximum capacity in kW
    c_fix["Boiler"] = 551.5  # Fix cost share Boiler in EUR
    part_load["Boiler"] = [0.3, 0.2]  # Minimal part load of boiler
    eta["Boiler"] = [0.7, 0.948]  # Thermal efficiency of a modern condensing boiler based on LHV
    f_serv["Boiler"] = 0.03  # Service costs based on total investment (c_fix + inv)
    t_life["Boiler"] = 18
    bounds["Boiler", "up"] = max(cap["Boiler"])  # Upper bound for boiler capacity in kW
    bounds["Boiler", "lo"] = 11  # Lower bound for boiler capacity in kW

    # CHP units

    # After continuous CHP

    inv["CHP"] = 2094  # Specific investment costs in Euro/kW for chp in Euro per kW electrical
    c_fix["CHP"] = 11291  # Fix cost share CHP in EUR
    cap["CHP"] = [5, 10]  # Not capacitites but rather 2 Upper bound for CHP capacity in kW to split it into  two
    # categories ... less than 5 and more than 5 KW ... so 2 decision variables (because of if cap<5 statement later
    # in the code)
    f_serv["CHP"] = 0.05  # Service costs based on investment
    t_life["CHP"] = 15
    bounds["CHP", "up"] = 10  # Upper bound for CHP capacity in kW
    eta["CHP", "total"] = 1.0  # Total efficiency of CHP unit using condensing technology
    eta["CHP", "el"] = 0.25  # Electrical efficiency CHP unit
    part_load["CHP"] = 0.5  # Minimal part load of CHP's with cap > 5000 W
    sigma = eta["CHP", "el"] / (eta["CHP", "total"] - eta["CHP", "el"])  # Power to heat ratio of CHP unit

    # Model in part load: P_el = alpha * Q_th_max + beta * Q_th + gamma [4, page 246, equation 4]
    params = {("CHP", "alpha"): -0.146,  # without unit
              ("CHP", "beta"): 0.66,   # without unit
              ("CHP", "gamma"): -2.620}  # kW

    # After continuous CHP
    #  moved it to directly after declaring the cap_design variables because for the model there still is no such thing
    #  as a cap_design for the CHP and it is no problem since the equation that uses q_CHP is not needed until way after
    #  declaring the cap_design variables

    # Before continuous CHP
    '''
    q_CHP = [(cap["CHP"][i] - params["CHP", "gamma"]) /
             (params["CHP", "alpha"] + params["CHP", "beta"])
             for i in range(len(cap["CHP"]))]
    '''
    # Photovoltaic system
    inv["PV"] = 1615.0      # Investment costs of PV system in EUR/kWp 
    f_serv["PV"] = 65.0   # Service costs for the PV system in EUR/(kWp*year)
    t_life["PV"] = 20
    bounds["PV", "up"] = 50  # Maximum area on the roof covered with PV-cells in m2
    bounds["PV", "low"] = 8  # entspricht 1kWp
    eta["PV"] = 0.14 * 0.85  # Efficiency of the photovoltaic cells (0.14) and further losses (0.85 -> converter, wiring...)
    pv_specific_capacity = 0.125  # Specific peak power per square meter in kWp/m2
    
    # Heat Pump
    # Investment costs for air-water heat pumps computed by linear regression fit(MA Rheinhardt)

    # After continuous HP

    inv["HP"] = 604.9  # Specific investment costs HP in EUR/kW
    f_serv["HP"] = 0.025  # Service costs based on investment
    t_life["HP"] = 18
    bounds["HP", "up"] = 25  # in kW
    c_fix["HP"] = 5316  # Fix cost share HP in EUR
    part_load["HP"] = 0.4  # Part load factor of HP
    t_flow = 35  # Set t_flow ONLY 35, 45 and 55 °C are valid!!
    t_return = 25


    # COP of HP at reference point A2W35
    if t_flow == 35:
        cop_ref = 3.8
    elif t_flow == 45:
        cop_ref = 2.96
    elif t_flow == 55:
        cop_ref = 2.44
    
    # Heat pump COP calculation:
    a_cop = 3.02367   # Parameter for COP curve
    b_cop = 126.96333  # Parameter for COP curve
    cop_lowest = (a_cop * T_ne + b_cop) / t_flow  # Calculation of COP at t_lowest
    cop = (a_cop * t_amb + b_cop) / t_flow  # Coefficient of performance of HP dependent on ambient temperature and t_flow

    # Electrical Heater
    inv["EH"] = 19  # Specific investment costs EH in EUR/kW_el
    c_fix["EH"] = 245  # Fix cost share electrical heater in EUR
    f_serv["EH"] = 0.0
    t_life["EH"] = 18
    bounds["EH", "up"] = 100  # dummy in kW
    eta["EH"] = 1     # Efficiency of electrical heater

    # Solar thermal collector
    inv["STC"] = 92.44  # EUR per m^2
    c_fix["STC"] = 596.24  # EUR fixed costs
    c_insurance = {"STC": 33.61}  # EUR per year (fixed costs)
    f_serv["STC"] = 0.015  # Service costs for solar thermal collectors. Based on initial investment
    t_life["STC"] = 30
    eta["STC", "optical"] = 0.8  # Optical efficiency, without unit
    eta["STC", "a1"] = 3.6390 / 1000  # kW per (m^2 K)
    eta["STC", "a2"] = 0.0168 / 1000  # kW per (m^2 K^2)

    t_stc_mean = (t_flow + t_return) / 2  # Mean temperature inside the collector
    
    eta["STC"] = (eta["STC", "optical"]
                  - eta["STC", "a1"] * (t_stc_mean - t_amb) / solar_rad
                  - eta["STC", "a2"] * ((t_stc_mean - t_amb) ** 2) / solar_rad)
    eta["STC"][np.isnan(eta["STC"])] = 0
    eta["STC"] = np.maximum(eta["STC"], 0)
    
    e_STC = np.zeros_like(eta["STC"])
    e_STC[eta["STC"] > 0] = 4.82 / 1000  # kW/m2 pumping electricity
    
    rval = {dev: (t_life[dev] - np.array([t_clc])) / t_life[dev] / (q ** t_clc) 
            for dev in t_life.keys()}  # Calculation of residual value [3]
    
    # Additional parameters
    number_devices = {dev: len(cap[dev]) for dev in cap.keys()}  # Battery, CHP, HP, TES, Boiler
    
    ###############################################################################
    ###############################################################################
    # setting up the model
    # try:
    # Create a new model
    model = gp.Model("Energy system design for a single building")
    
    # Create new variables
    # Costs
    c_total = model.addVar(vtype="C", lb=-np.inf, ub=np.inf)
    c_inv = {dev: model.addVar(vtype="C")
             for dev in t_life.keys()}                 
    c_serv = {dev: model.addVar(vtype="C", name="c_serv_"+dev)
              for dev in f_serv.keys()}
    c_dem = {dev: model.addVar(vtype="C", name="c_dem_"+dev)
             for dev in ["Boiler", "HP", "CHP", "Grid"]}
    c_met = model.addVar(vtype="C", name="c_metering")
    
    # Devices

    # My Addition
    # Kontinuierliche Variablen
    # After continuous TES
    # After continuous Boiler
    # After continuous HP
    # After continuous CHP

    cap_design = {dev: model.addVar(vtype="C", name="cap_" + dev)
                  for dev in ["EH", "Boiler", "Boiler_Conti", "TES", "TES_Conti", "PV", "Battery",
                              "HP", "CHP", "CHP0", "CHP1"]}
    initial = {dev: model.addVar(vtype="C", name="initial_" + dev) for dev in ["Battery"]}

    ### QUADRATIC CONSTRAINT PROBLEM DUE TO THE NEXT CONSTRAINT ###

    #q_CHP = [(cap_design["CHP" + str(i)] - params["CHP", "gamma"]) /
    #         (params["CHP", "alpha"] + params["CHP", "beta"])
    #         for i in range(len(cap["CHP"]))]
    #for i in range(len(cap["CHP"])):
    #    x = q_CHP[i]
#
    #    print(x)
    # Before continuous CHP
    '''
    was at another position ... immediatly after the params dictionnary was created for the CHP
    q_CHP = [(cap["CHP"][i] - params["CHP", "gamma"]) /
             (params["CHP", "alpha"] + params["CHP", "beta"])
             for i in range(len(cap["CHP"]))]
             
    when i thought only 1 continuous CHP
    q_CHP = (cap_design["CHP"] - params["CHP", "gamma"]) / (params["CHP", "alpha"] + params["CHP", "beta"])

    '''

    meter = {dev: model.addVar(vtype="C", name="meter_"+dev)
             for dev in ("gas",)}
    
    revenue = {key: model.addVar(vtype="C", name="revenue_"+key)
               for key in ("CHP", "feed-in")}
    
    subsidy = model.addVar(vtype="C", name="subsidy")  # Battery purchase
    # If battery is installed, the exported PV power has to be limited (60%)
    limit_pv_battery = model.addVar(vtype="C", name="limit_pv_battery")
    
    emission = model.addVar(vtype="C", name="CO2_emissions", lb=-gp.GRB.INFINITY)

    #My Addition
    # After continuous TES
    emissions_gas1 = model.addVar(vtype="C", name="CO2_emissions_gas", lb=-gp.GRB.INFINITY)
    emissions_PV1 = model.addVar(vtype="C", name="CO2_emissions_PV", lb=-gp.GRB.INFINITY)
    emissions_grid1 = model.addVar(vtype="C", name="CO2_emissions_Grid", lb=-gp.GRB.INFINITY)
    emissions_lca1 = model.addVar(vtype="C", name="CO2_emissions_Grid_LCA", lb=-gp.GRB.INFINITY)
    # Before continuous TES
    '''Did not exist'''

    energy = {}  # Energy consumption in Watt
    storage = {}  # Stored energy in Joule

    # ### Huge Change 2 to code old Start ###
    for t in range(timesteps):
        for d in range(number_days):
            timetag = "_" + str(d) + "_" + str(t)
            for dev in ("TES", "Battery"):
                storage[dev, d, t] = model.addVar(vtype="C",
                                                  name="storage_"+dev+timetag)
            for dev in ("CHP", "Boiler"):
                energy[dev, d, t] = model.addVar(vtype="C",
                                                 name="energy_"+dev+timetag)

            for dev in ("Boiler", "CHP", "HP", "EH", "STC"):
                heat[dev, d, t] = model.addVar(vtype="C",
                                               name="heat_"+dev+timetag)

            for dev in ("PV", "CHP", "HP", "EH", "Import", "STC"):
                power[dev, d, t] = model.addVar(vtype="C",
                                                name="power_"+dev+timetag)

            # My Addition
            # Kontinuierliche Variablen
            for dev in ("Battery",):
                power[dev, d, t, "charge"] = model.addVar(vtype="C",
                                                          name="power_"+dev+timetag+"_charge")
                power[dev, d, t, "discharge"] = model.addVar(vtype="C",
                                                             name="power_"+dev+timetag+"_discharge")
            # The way it was
            '''
            for dev in ("Battery",):
                power[dev, d, t, "charge"] = model.addVar(vtype="C",
                                                          name="power_"+dev+timetag+"_charge")
                power[dev, d, t, "discharge"] = model.addVar(vtype="C",
                                                             name="power_"+dev+timetag+"_discharge")
                for i in range(number_devices[dev]):
                    power[dev, d, t, i, "charge"] = model.addVar(vtype="C",
                                                                 name="power_"+dev+timetag+"_"+str(i)+"_charge")
                    power[dev, d, t, i, "discharge"] = model.addVar(vtype="C",
                                                                    name="power_"+dev+timetag+"_"+str(i)+"_discharge")
                    '''

            for dev in ("PV", "CHP"):
                power[dev, d, t, "use"] = model.addVar(vtype="C",
                                                       name="power_"+dev+timetag+"_use")
                power[dev, d, t, "sell"] = model.addVar(vtype="C",
                                                        name="power_"+dev+timetag+"_sell")

    # These variables are mainly used for linearizations
    q_max = {}  # Heat related
    p_max = {}  # Power related
    s_max = {}  # Storage related
    for t in range(timesteps):
        for d in range(number_days):
            timetag = "_" + str(d) + "_" + str(t)
            for dev in ("TES",):
                s_max[dev, d, t] = model.addVar(vtype="C", name="s_max_"+dev+timetag)
                
            for dev in ("HP",):
                q_max[dev, d, t] = model.addVar(vtype="C", name="q_max_"+dev+timetag)

             # After continuous CHP

            for dev in ("Boiler", "CHP"):
                for i in range(number_devices[dev]):
                    p_max[dev, d, t, i] = model.addVar(vtype="C",
                                                       name="p_max_"+dev+timetag+"_"+str(i))
                    q_max[dev, d, t, i] = model.addVar(vtype="C",
                                                       name="q_max_"+dev+timetag+"_"+str(i))
    
    x = {}  # Binary. 1 if device-type is installed, 0 otherwise
    for dev in number_devices.keys():
        for i in range(number_devices[dev]):
            x[dev, i] = model.addVar(vtype="B", name="x_"+dev+"_"+str(i))

    # My Addition
    # Kontinuierliche Variablen
    # After continuous HP
    # After continuous CHP

    for dev in ["PV", "STC", "EH", "Battery", "HP"]:
        x[dev] = model.addVar(vtype="B", name="x_"+dev)

    # The way it was
    # Before continuous HP
    # Before continuous CHP
    '''
    for dev in ["PV", "STC", "EH"]:
        x[dev] = model.addVar(vtype="B", name="x_"+dev)
        did not put CHP here anymore although it was when i had it as 1 continuous variable
        now it is already included in the upper part with range number_devices part
    '''
    ###
    ###
    y = {}  # Binary. 1 if device is activated, 0 otherwise

    # After Continuous HP
    # After continuous CHP

    for d in range(number_days):
        for t in range(timesteps):
            for dev in ("Boiler", "CHP"):
                for i in range(number_devices[dev]):
                    tag = dev + "_" + str(d) + "_" + str(t) + "_" + str(i)
                    y[dev, d, t, i] = model.addVar(vtype="B", name="y_" + tag)
            for dev in ("HP", ):
                tag = dev + "_" + str(d) + "_" + str(t)
                y[dev, d, t] = model.addVar(vtype="B", name="y_" + tag)


    area = {dev: model.addVar(vtype="C", name="a_"+dev)
            for dev in ("PV", "STC")}

    # 5R1C ISO 13790
    ###############
    # 2.2) Create new variables
    # 2.2.1) decision variables x for each device of building envelope (Binary)
    
    for i in range(len(U["window"])):
        x["window", i] = model.addVar(vtype="B", name="x_window_" + str(i))
    for drct2 in direction2:
        for j in range(len(U["opaque", drct2])):        
            x["opaque", drct2, j] = model.addVar(vtype="B",
                                                 name="x_opaque_" + drct2 + "_" + str(j))
    
    # binary variables for selection of building class (documentation z_cl)
                                     
    for l in range(len(f_class["Am"])):
        x["class", l] = model.addVar(vtype="B", name="x_class_" + str(l))
        
    # 2.2.2) Variables for linearization
    
    z_1 = {} 
    z_2 = {}  # z_2[d,t,drct2,j] = x["opaque",drtc2,j] * T_m
    z_3 = {}  # z_3[d,t,i] = x["window",i] * T_s
    z_4 = {} 
    z_5 = {}
    z_6 = {}  # z_6[d,t,l] = x["class",l] * T_m
    z_7 = {}  # z_7[d,t,l] = x["class",l] * T_s
    
    for d in range(number_days):
        for t in range(timesteps):
            timetag = str(d)+"_"+str(t)+"_"
            for drct2 in direction2:
                for j in range(len(U["opaque", drct2])):
                    # 2.2.2.1) linearization of z_2[d,t, drct2, j] = x["opaque", drtc2, j] * T_m
                    z_2[d, t, drct2, j] = model.addVar(vtype="C",
                                                       name="z_2_"+timetag+drct2+"_"+str(j))
            for i in range(len(U["window"])):
                # 2.2.2.2) linearization of z_3[d,t, i] = x["window", i] * T_s
                z_3[d, t, i] = model.addVar(vtype="C", name="z_3_"+timetag+str(i))
            for l in range(len(f_class["Am"])):
                # 2.2.2.3) linearization of z_6[d,t, l] = x["class", l] * T_m
                z_6[d, t, l] = model.addVar(vtype="C", name="z_6_"+timetag+str(l))
                # 2.2.2.4) linearization of z_7[d,t, l] = x["class", l] * T_s
                z_7[d, t, l] = model.addVar(vtype="C", name="z_7_"+timetag+str(l))
            
    for drct2 in direction2:
        for j in range(len(U["opaque", drct2])):
            for i in range(len(U["window"])):
                # 2.2.2.5) linearization of z_1[drct2, j, i] = x["window", i] * x["opaque", drct2, j]
                z_1[drct2, j, i] = model.addVar(vtype="B", 
                                                name="z_1_" + drct2 + "_" + str(j) + "_" + str(i))
            for l in range(len(f_class["Am"])):
                # 2.2.2.6) linearization of z_4[drct2, j, l] = x["class", l] * x["opaque", drtc2, j]
                z_4[drct2, j, l] = model.addVar(vtype="B", 
                                                name="z_4_" + drct2 + "_" + str(j) + "_" + str(l))
                
    for i in range(len(U["window"])):
        for l in range(len(f_class["Am"])):
            # 2.2.2.7) linearization of  z_5[i, l] = x["class", l] * x["window", i]
            z_5[i, l] = model.addVar(vtype="B", name="z_5_" + str(i) + "_" + str(l))
       
    # Continuous Variables for each timestep
    T_s = {}
    T_i = {}
    T_m = {}
    # T_m_init ={}
    Q_HC = {}  # in kW
    
    for d in range(number_days):
        for t in range(timesteps):
            timetag = str(d)+"_"+str(t)
        # Surface related temperature (DIN EN ISO 13790, section 7.2.2.1, page 34 text)
            T_s[d, t] = model.addVar(vtype="C", name="T_s_"+timetag)
        # Room temperature 
            T_i[d, t] = model.addVar(vtype="C", name="T_i_"+timetag)
        # Mass related temperature
            T_m[d, t] = model.addVar(vtype="C", name="T_m_"+timetag)
        # Heat demand
            Q_HC[d, t] = model.addVar(vtype="C", name="Q_HC_"+timetag,
                                      lb=0, ub=100)  # max. 100 kW
        
    # 5R1C ISO 13790 End
    ###################
     
    # Update model to integrate the new variables
    model.update()

    ###############################################################################
    ###############################################################################
    # Objective: Minimize investments, service, demand, metering costs (less 
    #   generated revenues)
    if options["opt_costs"]:
        model.setObjective(c_total, gp.GRB.MINIMIZE)
    else:
        model.setObjective(emission, gp.GRB.MINIMIZE)

    ###############################################################################
    ###############################################################################
    # Constraints
    
    # Costs
    model.addConstr(c_total == sum(c_inv[dev] for dev in c_inv.keys()) +
                    sum(c_serv[dev] for dev in c_serv.keys()) +
                    sum(c_dem[dev] for dev in c_dem.keys()) +
                    c_met - sum(revenue[key] for key in revenue.keys()))
    # Investment costs:
    # ------------------------------------------------------------------------------

    # After continuous TES
    # After continuous Boiler
    # After continuous HP
    # After continuous CHP
    for dev in ("TES", "Boiler"):
        model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * (x[dev, 0] * inv[dev][0]
                                                                        + cap_design[dev + "_Conti"] * inv[dev][1]
                                                                        + x[dev, 1] * c_fix[dev]))
    for dev in ("CHP", ):  # CHP is 2 distinct continuous variables now
        model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * ((cap_design[dev + "0"] * inv[dev]
                                                                         + x[dev, 0] * c_fix[dev])
                                                                        + (cap_design[dev + "1"] * inv[dev]
                                                                           + x[dev, 1] * c_fix[dev])))
    for dev in ("HP", ):
        model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * (inv[dev] * cap_design[dev]
                                                                        + c_fix[dev] * x[dev]))
    # Before continuous TES
    # Before continuous Boiler
    # Before continuous HP
    # Before continuous CHP
    '''
    for dev in ("HP", "TES", "Boiler", "CHP"):
        model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * sum(x[dev, i] * inv[dev][i]
                                                                           for i in range(number_devices[dev])))
     '''
    
    for dev in ("EH",):
        model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * (inv[dev] * cap_design[dev] + c_fix[dev] * x[dev]))
    
    dev = "PV"
    model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * inv[dev] * cap_design[dev])
    dev = "STC"
    model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * (inv[dev] * area[dev] + c_fix[dev] * x[dev]))
    dev = "Battery"

    # My Addition
    # Kontinuierliche Variablen

    model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * (inv[dev] * cap_design[dev]
                                                                    + c_fix[dev] * x[dev]) - 1 / t_clc * subsidy)

    # The way it was
    '''
    model.addConstr(c_inv[dev] == crf * (1 - rval[dev][0]) * tax * sum(x[dev, i] * inv[dev][i]
                                                                       for i in range(number_devices[dev])) - 1 / t_clc * subsidy)
    '''
    # 5R1C ISO 13790
    ###############
    
    for dev in direction2:
        model.addConstr(c_inv["opaque", dev] == crf * tax * 
                        sum((1 - rval["opaque", dev][i]) * x["opaque", dev, i] * A["opaque", dev] * inv["opaque", dev][i] 
                            for i in range(len(inv["opaque", dev]))))
    dev = "window"
    model.addConstr(c_inv[dev] == crf * tax * 
                    sum((1 - rval[dev][i]) * x[dev, i] * A[dev] * inv[dev][i]
                        for i in range(len(inv[dev]))))
    
    for dev in ("intWall", "ceiling", "intFloor"):
        model.addConstr(c_inv["opaque", dev] == 0.5 * crf * tax * 
                        sum((1 - rval["opaque", dev][i]) * x["opaque", "wall", i] * A["opaque", dev] * inv["opaque", dev][i]
                            for i in range(len(inv["opaque", dev]))))
    # 5R1C ISO 13790 End
    ###################
    # Invest battery    
    # Battery subsidies
    dev = "Battery"

    # My Addition
    # Kontinuierliche Variablen

    model.addConstr(subsidy <= sub_battery["PV", "sub"] *
                    ((inv[dev] * cap_design[dev] + c_fix[dev] * x[dev]) +
                    (inv["PV"] - sub_battery["PV", "investment"]) * cap_design["PV"]))
    model.addConstr(subsidy <= sub_battery["PV", "sub"] * sub_battery["PV", "bound"] * cap_design["PV"])
    model.addConstr(subsidy <= sub_battery["PV", "sub"] * sub_battery["PV", "bound"] * x[dev])

    # The way it was
    '''
    model.addConstr(subsidy <= sub_battery["PV", "sub"] *
                    (sum(x[dev, i] * inv[dev][i] for i in range(number_devices[dev])) +
                     (inv["PV"] - sub_battery["PV", "investment"]) * cap_design["PV"]))
    model.addConstr(subsidy <= sub_battery["PV", "sub"] * sub_battery["PV", "bound"] * cap_design["PV"])
    model.addConstr(subsidy <= sub_battery["PV", "sub"] * sub_battery["PV", "bound"] *
                    sum(x[dev, i] for i in range(number_devices[dev])))
    '''
    #
    # Service and operation costs
    # ------------------------------------------------------------------------------
    dev = "PV"

    model.addConstr(c_serv[dev] == crf * b["infl"] * tax * f_serv[dev] * cap_design[dev])
    
    dev = "STC"
    model.addConstr(c_serv[dev] == crf * b["infl"] * tax *
                    (f_serv[dev] * (inv[dev] * area[dev] + c_fix[dev] * x[dev]) + c_insurance["STC"] * x["STC"]))

    # After continuous Boiler
    # After continuous HP
    # After continuous CHP

    for dev in ("CHP", ):
        model.addConstr(c_serv[dev] == crf * b["infl"] * tax * f_serv[dev] * ((cap_design[dev + "0"] * inv[dev]
                                                                               + x[dev, 0] * c_fix[dev])
                                                                              + (cap_design[dev + "1"] * inv[dev]
                                                                                 + x[dev, 1] * c_fix[dev])))

    for dev in ("HP", ):
        model.addConstr(c_serv[dev] == crf * b["infl"] * tax * f_serv[dev]
                        * (cap_design[dev] * inv[dev] + x[dev] * c_fix[dev]))

    for dev in ("Boiler",):
        model.addConstr(c_serv[dev] == crf * b["infl"] * tax * f_serv[dev]
                        * (x[dev, 0] * inv[dev][0] + cap_design[dev + "_Conti"] * inv[dev][1] + x[dev, 1] * c_fix[dev]))

    # Before continuous Boiler
    # Before continuous HP
    # Before continuous CHP
    '''
    for dev in ("CHP", "HP", "Boiler"):
        model.addConstr(c_serv[dev] == crf * b["infl"] * tax * f_serv[dev] *
                        sum(x[dev, i] * inv[dev][i] for i in range(number_devices[dev])))
    '''

    # Demand related costs
    # ------------------------------------------------------------------------------

    # ### Huge Change extra to code Old Start ###
    for dev in ("Boiler", "CHP"):
       model.addConstr(c_dem[dev] == crf * b["gas"] * price["gas", dev] * dt *
                       sum(weight_days[d] *
                           sum(energy[dev, d, t] for t in range(timesteps))
                           for d in range(number_days)))
    model.addConstr(c_dem["Grid"] == crf * b["el"] * price["el"] * dt *
                   sum(weight_days[d] *
                       sum(power["Import", d, t] for t in range(timesteps))
                       for d in range(number_days)))
    # ### Huge Change extra to code Old End  ###
    
    # Costs for metering
    # ------------------------------------------------------------------------------
    model.addConstr(c_met == crf * b["infl"] * (c_meter["gas"] * meter["gas"]))
    
    # Revenue
    # ------------------------------------------------------------------------------

    # ### Huge Change extra to code Old Start  ###
    model.addConstr(revenue["CHP"] == crf * b["eex"] * price["CHP"] * dt *
                    sum(weight_days[d] *
                        sum(power["CHP", d, t] for t in range(timesteps))
                        for d in range(number_days)))
    model.addConstr(revenue["feed-in"] == crf * b["eex"] * dt *
                    sum(weight_days[d] *
                        sum(power["CHP", d, t, "sell"] * price["CHP", "sell"][d, t] +
                            power["PV", d, t, "sell"] * price["PV", "sell"]
                            for t in range(timesteps))
                        for d in range(number_days)))
    # ### Huge Change extra to code Old End  ###

       
    # Metering constraints
    # After continuous CHP
    for dev in ("Boiler", "CHP" ):
        model.addConstr(meter["gas"] >= sum(x[dev, i] for i in range(number_devices[dev])))


    # CO2 emissions

    # ### Huge Change Extra to code Old Start ###
    emissions_gas = emi["gas"] * sum(weight_days[d] * dt *
                                     sum(sum(energy[dev, d, t]
                                             for dev in ("Boiler", "CHP"))
                                         for t in range(timesteps))
                                     for d in range(number_days))
    emissions_PV = emi["PV"] * sum(weight_days[d] * dt *
                                   sum(power["PV", d, t] for t in range(timesteps))
                                   for d in range(number_days))
    # ### Huge Change Extra to code Old End ###


    # ### Huge Change Extra to code old Start ###
    months_start_in_hours = []
    bb = 0
    for i in weight_days:
        months_start_in_hours.append(
            bb + 1)  # I want to switch no on the last hour of the months but on the first of the next ... this makes sense the way i used this list in the following loop
        bb += i * 24
    emissions_grid = 0
    dd = -1
    tt = 0
    # ######### LOOP FOR CO2(t) 2018 ############
    for i in range(1, 8761):
        if i in months_start_in_hours:
            dd += 1
            # print("using type day" + str(d))
            if tt == 23:
                emissions_grid += power["Import", dd, tt] * ws1.cell(row=i+1, column=5).value  # 0.423813555  # ws1.cell(row=i+1, column=5).value  # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value # What i need to do is replace emi el with ws1.cell(row=i+1, column=5).value ... so that i use the correct CO2(t)
                tt = 0
            else:
                emissions_grid += power["Import", dd, tt] * ws1.cell(row=i+1, column=5).value  # 0.423813555  # ws1.cell(row=i+1, column=5).value  # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value # if i want to test with original emissions again i would have to replace ws1.cell(row=i+1, column=5).value again with emi["el"
                tt += 1
        else:
            if tt == 23:
                emissions_grid += power["Import", dd, tt] * ws1.cell(row=i+1, column=5).value  # 0.423813555  # ws1.cell(row=i+1, column=5).value  # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value
                tt = 0
            else:
                emissions_grid += power["Import", dd, tt] * ws1.cell(row=i+1, column=5).value  # 0.423813555  # ws1.cell(row=i+1, column=5).value  # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value
                tt += 1
    # ### Huge Change Extra to code old End ###


    # My Additon
    # ######### LOOP FOR CO2(t) 2030 ############

    #for i in range(1, 8761):
    #    if i in months_start_in_hours:
    #        dd += 1
    #        print("using type day" + str(d))
    #        if tt == 23:
    #            emissions_grid += power["Import", dd, tt] * ws11.cell(row=i+1, column=5).value # 0.480657476 #  ws11.cell(row=i+1, column=5).value #    # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value # What i need to do is replace emi el with ws1.cell(row=i+1, column=5).value ... so that i use the correct CO2(t)
    #            tt = 0
    #        else:
    #            emissions_grid += power["Import", dd, tt] * ws11.cell(row=i+1, column=5).value # 0.480657476 #  ws11.cell(row=i+1, column=5).value #    # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value # if i want to test with original emissions again i would have to replace ws1.cell(row=i+1, column=5).value again with emi["el"
    #            tt += 1
    #    else:
    #        if tt == 23:
    #            emissions_grid += power["Import", dd, tt] * ws11.cell(row=i+1, column=5).value # 0.480657476 #  ws11.cell(row=i+1, column=5).value #    # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value
    #            tt = 0
    #        else:
    #            emissions_grid += power["Import", dd, tt] * ws11.cell(row=i+1, column=5).value # 0.480657476 #  ws11.cell(row=i+1, column=5).value #    # i replaced the variable factor with a constant value that represents the average ... i just need to replace it by what is directly after the comment # ws1.cell(row=i+1, column=5).value
    #            tt += 1

    # My Addition
    # Emissions due to the components them selves (LCA)

    emi_lca_materials = {"ferroconcrete": 237.34,  # kg CO2 Äqu/ m3
                         "polystyrene": 55.27,  # kg CO2 Äqu/ m3
                         "concrete layer for outer walls": 242,  # kg CO2 Äqu/ m3
                         "concrete layer": 242,# kg CO2 Äqu/ m3
                         "gypsum plasterboard": 1.5745,  # kg CO2 Äqu/ m2
                         "sandwich panel": 10.73,  # kg CO2 Äqu/ m3
                         "core insulation": 55.27,  # kg CO2 Äqu/ m3
                         "air gap": 0,  # kg CO2 Äqu/ m3
                         "lime stone": 303.4,  # kg CO2 Äqu/ m3
                         "Plastering": 718.4,  # kg CO2 Äqu/ m3
                         "Lime Plaster": 405.1,  # kg CO2 Äqu/ m3
                         "concrete for internal wall": 69.2,  # # kg CO2 Äqu/ m2
                         "concrete for ceiling": 242,  # kg CO2 Äqu/ m3
                         "cement screed": 367.2,  # kg CO2 Äqu/ m3
                         "concrete for ground floor": 237.34,  # kg CO2 Äqu/ m3
                         "foam glas": 158.4,  # kg CO2 Äqu/ m3
                         "gravel fill": 32.42,  # kg CO2 Äqu/ m3
                         "insulation board": 266,  # kg CO2 Äqu/ m3
                         "Wooden Frame": -0.26,  # kg CO2 Äqu/ m
                         "double glazing": 26.7,  # kg CO2 Äqu/ m2
                         "Plastic frame":  8.17,  # kg CO2 Äqu/ m
                         "insulation glazing": 9.87,  # kg CO2 Äqu/ m2
                         "double insulation glazing": 37,  # kg CO2 Äqu/ m2
                         "Triple insulation glazing": 50.16  # kg CO2 Äqu/ m2
                         }

    Area_outside_walls = 4 * 42.25  # m2
    Area_int_walls = 375  # m2
    Area_roof = 99.75 #m2
    Area_floor = 99.75 #m2
    A_ceiling = 75  #m2
    A_int_floor = 75  #m2
    Area_windows = 4 * 7.5 #m2


    ow_0 = 0.175 * emi_lca_materials["ferroconcrete"] + 0.03 * emi_lca_materials["polystyrene"] + 0.08 * emi_lca_materials["concrete layer for outer walls"]
    ow_1 = ow_0 + 0.0125 * emi_lca_materials["gypsum plasterboard"] + 0.04 * emi_lca_materials["sandwich panel"]
    ow_2 = ow_0 + 0.008 * emi_lca_materials["Plastering"] + 0.15 * emi_lca_materials["core insulation"] + 0.01 * emi_lca_materials["air gap"] + 0.115 * emi_lca_materials["lime stone"]
    ow_3 = ow_2 + 0.07 * emi_lca_materials["core insulation"]
    iw_0 = 0.01 * emi_lca_materials["Lime Plaster"] + 0.10 * emi_lca_materials["concrete for internal wall"] + 0.01 * emi_lca_materials["Lime Plaster"]
    ceil_0 = 0.16 * emi_lca_materials["concrete for ceiling"] + 0.06 * emi_lca_materials["polystyrene"] + 0.04 * emi_lca_materials["cement screed"]
    flr_0 = 0.04 * emi_lca_materials["cement screed"] + 0.06 * emi_lca_materials["polystyrene"] + 0.16 * emi_lca_materials["concrete layer"]
    gf_0 = 0.04 * emi_lca_materials["cement screed"] + 0.03 * emi_lca_materials["polystyrene"] + 0.15 * emi_lca_materials["concrete for ground floor"]
    rf_0 = 0.15 * emi_lca_materials["ferroconcrete"] + 0.07 * emi_lca_materials["foam glas"] + 0.03 * emi_lca_materials["gravel fill"]
    rf_1 = 0.15 * emi_lca_materials["ferroconcrete"] + 0.09 * emi_lca_materials["foam glas"] + 0.03 * emi_lca_materials["gravel fill"]
    rf_2 = 0.15 * emi_lca_materials["ferroconcrete"] + 0.24 * emi_lca_materials["insulation board"] + 0.03 * emi_lca_materials["gravel fill"]
    rf_3 = 0.15 * emi_lca_materials["ferroconcrete"] + 0.36 * emi_lca_materials["insulation board"] + 0.03 * emi_lca_materials["gravel fill"]
    win_0 = emi_lca_materials["Wooden Frame"] + Area_windows * emi_lca_materials["double glazing"]  # Total window area = 30 m squared from Building.py line 20
    win_1 = emi_lca_materials["Plastic frame"] + Area_windows * emi_lca_materials["insulation glazing"]
    win_2 = emi_lca_materials["Plastic frame"] + Area_windows * emi_lca_materials["double insulation glazing"]
    win_3 = emi_lca_materials["Plastic frame"] + Area_windows * emi_lca_materials["Triple insulation glazing"]
    factor_STC = 104.3  # kg CO2 Äqu/ m2
    factor_EH = 0.71  # kg CO2 Äqu/ kw
    factor_PV = 304  # kg CO2 Äqu/ m2
    factor_Battery = 243.9 #  243.9 #0  # kg CO2 Äqu/ kwh
    factor_TES = 3.60  # kg CO2 Äqu/ m3
    factor_Boiler= 21.40  # kg CO2 Äqu/ kw
    factor_HP= 45.30  # kg CO2 Äqu/ kw
    factor_CHP = 21.40  # kg CO2 Äqu/ kwel

    emi_lca_components = {"Outer Wall 0": 0 * (Area_outside_walls * ow_0) / 40,  # Life time of walls = 40 years ... already installed so 0 emissions
                          "Outer Wall 1": (Area_outside_walls * ow_1) / 40,  # Life time of walls = 40 years
                          "Outer Wall 2": (Area_outside_walls * ow_2) / 40,  # Life time of walls = 40 years
                          "Outer Wall 3": (Area_outside_walls * ow_3) / 40,  # Life time of walls = 40 years
                          # "Internal Wall": iw_0, # always there ... not an option to be added or saniert
                          # "Ceiling": ceil_0, # always there ... not an option to be added or saniert
                          "Floor": 0 * (Area_floor * flr_0) / 40,   # Life time of walls = 40 years ... already installed so 0 emissions
                          # "Ground Floor": # always there ... not an option to be added or saniert
                          "Roof 0": 0 * (Area_roof * rf_0) / 40,  # Life time of roof = 40 years  ... already installed so 0 emissions
                          "Roof 1": (Area_roof * rf_1) / 40,  # Life time of roof = 40 years
                          "Roof 2": (Area_roof * rf_2) / 40,  # Life time of roof = 40 years
                          "Roof 3": (Area_roof * rf_3) / 40,  # Life time of roof = 40 years
                          "Windows 0": 0 * (win_0) / 40,  # Life time of windows = 40 years ... already installed so 0 emissions
                          "Windows 1": (win_1) / 40,  # Life time of windows = 40 years
                          "Windows 2": (win_2) / 40,  # Life time of windows = 40 years
                          "Windows 3": (win_3) / 40,  # Life time of windows = 40 years
                          "Battery": (cap_design["Battery"] * factor_Battery) / 13.7,  #
                          # "Battery 0": 0, #Add the correct ones accounting for capacity and lifetime
                          # "Battery 1": 0, #Add the correct ones accounting for capacity and lifetime
                          # "Battery 2": 0, #Add the correct ones accounting for capacity and lifetime
                          # "Battery 3": 0, #Add the correct ones accounting for capacity and lifetime
                          "TES 0": 0,  # Already installed so 0 emissions
                          "TES 1": (cap_design["TES_Conti"] * factor_TES) / 20,  # Life time of TES = 20 years
                          # "TES 2": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "TES 3": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "TES 4": 0,  # Add the correct ones accounting for capacity and lifetime
                          "Boiler 0": 0,  # Already installed so 0 emissions
                          "Boiler 1": (cap_design["Boiler_Conti"] * factor_Boiler) / 18,  # Life time of
                          # "Boiler 2": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "Boiler 3": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "Boiler 4": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "Boiler 5": 0,  #Add the correct ones accounting for capacity and lifetime
                          "CHP 0": (cap_design["CHP0"] * factor_CHP) / 15,  # Life time of CHP = 15 years
                          "CHP 1": (cap_design["CHP1"] * factor_CHP) / 15,  # Life time of CHP = 15 years
                          # "CHP 2": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "CHP 3": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "CHP 4": 0,  #Add the correct ones accounting for capacity and lifetime
                          "PV": (area["PV"] * factor_PV) / 20,  # Life time of PV = 20
                          "HP": (cap_design["HP"] * factor_HP) / 18,  # Life time of HP = 18
                          # "HP 1": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "HP 2": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "HP 3": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "HP 4": 0,  #Add the correct ones accounting for capacity and lifetime
                          # "HP 5": 0,  #Add the correct ones accounting for capacity and lifetime
                          "EH": (cap_design["EH"] * factor_EH) / 18,  # Life time of EH = 18
                          "STC": (area["STC"] * factor_STC) / 30,  #Add the correct ones accounting for capacity and lifetime
                          }


    decision_variable = {"Outer Wall 0": x["opaque", "wall", 0],
                         "Outer Wall 1": x["opaque", "wall", 1],
                         "Outer Wall 2": x["opaque", "wall", 2],
                         "Outer Wall 3": x["opaque", "wall", 3],
                         # "Internal Wall": 1 , # always there ... not an option to be added or saniert
                         # "Ceiling": 1 , # always there ... not an option to be added or saniert
                         "Floor": x["opaque", "floor", 0],
                         # "Ground Floor": gf_0, # always there ... not an option to be added or saniert
                         "Roof 0": x["opaque", "roof", 0],
                         "Roof 1": x["opaque", "roof", 1],
                         "Roof 2": x["opaque", "roof", 2],
                         "Roof 3": x["opaque", "roof", 3],
                         "Windows 0": x["window", 0],
                         "Windows 1": x["window", 1],
                         "Windows 2": x["window", 2],
                         "Windows 3": x["window", 3],
                         "Battery": x["Battery"],
                         # "Battery 0": x["Battery", 0],
                         # "Battery 1": x["Battery", 1],
                         # "Battery 2": x["Battery", 2],
                         # "Battery 3": x["Battery", 3],
                         "TES 0": x["TES", 0],
                         "TES 1": x["TES", 1],
                         # "TES 2": x["TES", 2],
                         # "TES 3": x["TES", 3],
                         # "TES 4": x["TES", 4],
                         "Boiler 0": x["Boiler", 0],
                         "Boiler 1": x["Boiler", 1],
                         # "Boiler 2": x["Boiler", 2],
                         # "Boiler 3": x["Boiler", 3],
                         # "Boiler 4": x["Boiler", 4],
                         # "Boiler 5": x["Boiler", 5],
                         "CHP 0": x["CHP", 0],
                         "CHP 1": x["CHP", 1],
                         #"CHP 2": x["CHP", 2],
                         #"CHP 3": x["CHP", 3],
                         #"CHP 4": x["CHP", 4],
                         "PV": x["PV"],
                         "HP": x["HP"],
                         # "HP 1": x["HP", 1],
                         # "HP 2": x["HP", 2],
                         # "HP 3": x["HP", 3],
                         # "HP 4": x["HP", 4],
                         # "HP 5": x["HP", 5],
                         "EH": x["EH"],
                         "STC": x["STC"],
                         }
    emissions_lca = 0

    for component in emi_lca_components.keys():
        if component in ["EH", "Battery", "STC", "PV", "TES 1", "Boiler 1", "CHP 0", "CHP 1", "HP"]:
            emissions_lca += (emi_lca_components[component])  # No need to convert from kg to tonnes here because it is done at the end with the .X method/ 1000  # Convert from kgs to Tonnes
        else:
            emissions_lca += (decision_variable[component] * emi_lca_components[component]) # No need to convert from kg to tonnes here because it is done at the end with the .X method/ 1000  # Convert from kgs to Tonnes


    model.addConstr(emission == emissions_gas + emissions_PV + emissions_grid + emissions_lca)

    # After continuous TES
    model.addConstr(emissions_gas1 == emissions_gas)
    model.addConstr(emissions_PV1 == emissions_PV)
    model.addConstr(emissions_grid1 == emissions_grid)
    model.addConstr(emissions_lca1 == emissions_lca)

    #Before continuous TES
    '''Did not Exist'''

    #My Addition
    #Here i added te emissions_lca in the end of the calculation of emissions
    #model.addConstr(emission == emissions_gas + emissions_PV + emissions_grid + emissions_lca)
    model.addConstr(0.001 * emission <= emission_max)
    
    # Technology related and logical constraints
    # ------------------------------------------------------------------------------
    # Operation and design of devices
    # Design of devices (constraints without time dependency)

    # Boiler

    # After continuous Boiler
    # included with the calculation of cap_design TES in the lines below in the equations of thermal storage

    # Before continuous TES
    '''
    dev = "Boiler"
    model.addConstr(cap_design[dev] == sum(x[dev, i] * cap[dev][i]
                                           for i in range(number_devices[dev])))
    model.addConstr(sum(x[dev, i] for i in range(number_devices[dev])) <= 1)
    '''
    # Electrical heater
    model.addConstr(cap_design["EH"] <= bounds["EH", "up"] * x["EH"])

    # After continuous TES
    # After continuous Boiler
    # After continuous HP
    # After continuous CHP
    # To remove illogical negative capacity
    model.addConstr(cap_design["EH"] >= 0)
    model.addConstr(cap_design["HP"] >= 0)
    model.addConstr(cap_design["CHP"] >= 0)
    model.addConstr(cap_design["Boiler" + "_Conti"] >= 0)
    model.addConstr(cap_design["TES" + "_Conti"] >= 0)
    model.addConstr(area["STC"] >= 0)
    model.addConstr(cap_design["Boiler" + "_Conti"] <= cap["Boiler"][1] * x["Boiler", 1])

    # Before continuous TES
    # Before continuous HP
    # Before continuous CHP
    ''' It did not exist before '''

    # My Addition
    # Kontinuierliche Variablen
    model.addConstr(cap_design["Battery"] <= bounds["Battery", "up"] * x["Battery"])

    # After continuous CHP
    model.addConstr(cap_design["CHP0"] <= cap["CHP"][0] * x["CHP", 0])
    model.addConstr(cap_design["CHP1"] <= cap["CHP"][1] * x["CHP", 1])
    model.addConstr(cap_design["CHP" + "0"] >= 0)
    model.addConstr(cap_design["CHP" + "1"] >= ((cap["CHP"][0] + 0.001) * x["CHP", 1]))
    model.addConstr(cap_design["CHP"] == cap_design["CHP0"] + cap_design["CHP1"])

    # CHP1 is greater than 5 as soon as it is chosen and less than 10 ... so greater than max CHP0 and less than max CHP1

    # The way it was
    # Before continuous CHP
    ''' It did not exist before '''

    # After continuous HP
    model.addConstr(cap_design["HP"] <= bounds["HP", "up"] * x["HP"])

    # The way it was
    # Before continuous HP
    ''' It did not exist before '''

    # After continuous TES
    # After continuous Boiler
    for dev in ("TES", "Boiler"):
        model.addConstr(cap_design[dev + "_Conti"] <= bounds[dev, "up"] * x[dev, 1])
        model.addConstr(cap_design[dev + "_Conti"] >= bounds[dev, "lo"] * x[dev, 1])

    # Before continuous TES
    ''' It did not exist before '''
    
    # Thermal storage

    # After continuous TES
    # After continuous Boiler
    for dev in ("TES", "Boiler"):
        model.addConstr(cap_design[dev] == x[dev, 0] * cap[dev][0] + cap_design[dev+"_Conti"])
        model.addConstr(sum(x[dev, i] for i in range(number_devices["TES"])) <= 1)

    # Before continuous TES
    '''
    model.addConstr(cap_design["TES"] == sum(x["TES", i] * cap["TES"][i]
                                             for i in range(number_devices["TES"])))
    model.addConstr(sum(x["TES", i] for i in range(number_devices["TES"])) <= 1)
    '''

    # Battery
    dev = "Battery"

    # My Addition
    # Kontinuierliche Variablen

    sum_x_bat = x[dev]
    model.addConstr(sum_x_bat <= 1)

    # The way it was
    '''
    # Introduce abbreviation for sum(i, x["Battery",i])
    sum_x_bat = sum(x[dev, i] for i in range(number_devices[dev]))
    model.addConstr(sum_x_bat <= 1)
    '''

    # Linearization of limit_pv_battery = x["Battery",i]*cap_design["PV"]
    model.addConstr(limit_pv_battery <= sum_x_bat * min(bounds["PV", "up"], A["opaque", "floor"]) * pv_specific_capacity)
    model.addConstr(cap_design["PV"] - limit_pv_battery >= 0)
    model.addConstr(cap_design["PV"] - limit_pv_battery <= 
                    (1 - sum_x_bat) * min(bounds["PV", "up"], A["opaque", "floor"]) * pv_specific_capacity)
    
    # PV and STC
    for dev in area.keys():
        model.addConstr(area[dev] <= min(bounds["PV", "up"], A["opaque", "floor"]) * x[dev])
    model.addConstr(area["PV"] + area["STC"] <= min(bounds["PV", "up"], A["opaque", "floor"]))
    model.addConstr(area["PV"] * pv_specific_capacity == cap_design["PV"])
    model.addConstr(area["PV"] >= bounds["PV", "low"] * x["PV"])
    
    ###
    # 5R1C ISO 13790
    ###############
    
    # Nominal heat load
    # Q_nHC = Q_T,i + Q_v_i + Q_RH,i (REG1 VL04)
    # Q_RH,i = A_f * f_RH not considered!!!
    Q_nHC = (sum(A["opaque", "wall"] * U["opaque", "wall"][i] * x["opaque", "wall", i] 
                 for i in range(len(U["opaque", "wall"]))) + 
             sum(A["window"] * U["window"][i] * x["window", i] 
                 for i in range(len(U["window"]))) + 
             sum(A["opaque", "roof"] * U["opaque", "roof"][i] * x["opaque", "roof", i] 
                 for i in range(len(U["opaque", "roof"]))) + 
             sum(A["opaque", "floor"] * U["opaque", "floor"][i] * x["opaque", "floor", i]
                 for i in range(len(U["opaque", "floor"]))) * f_g1 * f_g2 * G_w
             + ventilationRate * c_p_air * rho_air * V / 3600) * (T_set_min - T_ne)

    # After continuous HP
    # After continuous CHP


    model.addConstr(((cap_design["CHP"] / sigma)) +
                    cap_design["Boiler"] + cap_design["EH"] * eta["EH"] +
                    cop_lowest / cop_ref * cap_design["HP"] >= 0.001 * Q_nHC)

     ### QUADRATIC CONSTRAINT PROBLEM DUE TO THE NEXT CONSTRAINT ###
    # Had a problem with because it used to be like this which made it non linear

    # model.addConstr(((cap_design["CHP"] / sigma) * (x["CHP", 0] + x["CHP", 1])) +
    #                 cap_design["Boiler"] + cap_design["EH"] * eta["EH"] +
    #                 cop_lowest / cop_ref * cap_design["HP"] >= 0.001 * Q_nHC)

    # Before continuous HP
    # Before continuous CHP
    '''
    model.addConstr(sum(cap["CHP"][i] / sigma * x["CHP", i] for i in range(number_devices["CHP"])) +
                    cap_design["Boiler"] + cap_design["EH"] * eta["EH"] + 
                    cop_lowest / cop_ref * (sum(cap["HP"][i] * x["HP", i] for i in range(number_devices["HP"]))) >=
                    0.001 * Q_nHC)
    '''
    # 5R1C ISO 13790 End
    ###################
    ###
    ###
    # HP, CHP

    # After continuous HP
    # After continuous CHP
    for dev in ("HP", ):
        # I think this statement is redundant since x is a gurobi binary variable by definition
        # So this means it is always either 0 or 1
        # But it does not hurt to add it ... it still is true
        model.addConstr(x[dev] <= 1)

    for dev in ("CHP", ):
        model.addConstr(sum(x[dev, i] for i in range(number_devices[dev])) <= 1)

    # Before continuous HP
    # Before continuous CHP
    '''
    for dev in ("HP", "CHP"):
        model.addConstr(sum(x[dev, i] for i in range(number_devices[dev])) <= 1)
    '''

    q_CHP = [(cap_design["CHP" + str(i)] - (params["CHP", "gamma"] * x["CHP",i])) /
             (params["CHP", "alpha"] + params["CHP", "beta"])
             for i in range(len(cap["CHP"]))]
    # here i multiplied params gamma by x chp in order to make the numerator 0 if cap design is 0 because of the way this factor is used later in the calculation of q_maxCHP1
    # and the equation is still linear because params gamma is a constant


    for i in range(len(cap["CHP"])):
        xx = q_CHP[i]

        print(xx)


    # # ### HUGE CHANGE old Start###
    for d in range(number_days):
        # Storage cycling
        model.addConstr(storage["TES", d, timesteps-1] == init["TES"])
        dev = "Battery"

        # My Addition
        # Kontinuierliche Variablen

        # test change here
        #model.addConstr(initial[dev] >= 0.2 * cap_design[dev])  # The initial state of charge of the battery

        '''uncomment the next line only and delete the previous one to undo the change'''
        model.addConstr(initial[dev] == 0.5 * cap_design[dev])  # The initial state of charge of the battery
        # This constraint of initial dev can be outside of the loop since it does not change with d
        # It is not contributing to wrong results in the loop ... it is just that
        # The constraint is being reinitialized 12 times if gurobi does not skip it automatically
        # It should still give the same result if it was outside the loop and would be used the same way in the
        # next Constraint for storage dev,d,...
        model.addConstr(storage[dev, d, timesteps - 1] == initial[dev])

        # The way it was
        '''
        model.addConstr(storage[dev, d, timesteps-1] == sum(init[dev][i] * x[dev, i]
                                                            for i in range(number_devices[dev])))
        '''

        for t in range(timesteps):
            # Electricity balance (generation = consumption)

            # My Addition
            # Kontinuierliche Variablen
            # HERE
            #model.addConstr(power["Import", d, t] +
            #                sum(power[generator, d, t, "use"] for generator in ("CHP", "PV")) +
            #                power["Battery", d, t, "discharge"] * (1/eta["Battery", "discharge"]) ==
            #                sum(power[consumer, d, t] for consumer in ("HP", "EH", "STC")) +
            #                power["House"][d, t] +
            #                power["Battery", d, t, "charge"] * eta["Battery", "charge"])

            model.addConstr(power["Import", d, t] +
                            sum(power[generator, d, t, "use"] for generator in ("CHP", "PV")) +
                            power["Battery", d, t, "discharge"]  ==
                            sum(power[consumer, d, t] for consumer in ("HP", "EH", "STC")) +
                            power["House"][d, t] +
                            power["Battery", d, t, "charge"] )

            # TO TEST WITH 1 KWH and 100 KWH constant  consumption ... replace Power "House" By 1 kwh

            # model.addConstr(power["Import", d, t] +
            #                sum(power[generator, d, t, "use"] for generator in ("CHP", "PV")) +
            #                power["Battery", d, t, "discharge"] * (1 / eta["Battery", "discharge"]) ==
            #                sum(power[consumer, d, t] for consumer in ("HP", "EH", "STC")) +
            #                20 +
            #                power["Battery", d, t, "charge"] * eta["Battery", "charge"])

            # The way it was
            '''
            model.addConstr(power["Import", d, t] +
                            sum(power[generator, d, t, "use"] for generator in ("CHP", "PV")) +
                            sum(power["Battery", d, t, i, "discharge"] for i in range(number_devices["Battery"])) ==
                            sum(power[consumer, d, t] for consumer in ("HP", "EH", "STC")) +
                            power["House"][d, t] +
                            sum(power["Battery", d, t, i, "charge"] for i in range(number_devices["Battery"])))
            '''

            # Thermal storage
            # Energy balance
            if t > 0:
                storage_previous = storage["TES", d, t-1]
                battery_previous = storage["Battery", d, t-1]
            else:
                storage_previous = init["TES"]
                dev = "Battery"

                # My Addition
                # Kontinuierliche Variablen

                battery_previous = initial[dev]

                # The way it was
                '''
                battery_previous = sum(init[dev][i] * x[dev, i] for i in range(number_devices[dev]))
                '''
            model.addConstr(storage["TES", d, t] == (1-sto_loss) * storage_previous +
                            dt * (sum(heat[dev, d, t] for dev in ("Boiler", "CHP", "HP", "EH", "STC"))) -
                            dt * (Q_HC[d, t]) - dt * heat["DHW"][d, t])

            # Design restriction

            # After continuous TES
            model.addConstr(storage["TES", d, t] <= cap_design["TES"] * rho_water * cp_water * deltaT["TES"])
            # Before continuous TES
            '''
            model.addConstr(storage["TES", d, t] <= cap_design["TES"] * rho_water * cp_water * deltaT["TES"])
            '''
            # Battery storage
            dev = "Battery"

            # My Addition
            # Kontinuierliche Variablen

            model.addConstr(storage[dev, d, t] == battery_previous +
                            dt * ((power[dev, d, t, "charge"] * eta["Battery", "charge"]) - (power[dev, d, t, "discharge"] * (1 / eta["Battery", "discharge"]))))
            model.addConstr(storage[dev, d, t] >= (1 - bounds["Battery", "DOD"]) * cap_design["Battery"])
            model.addConstr(storage[dev, d, t] <= cap_design["Battery"])
            model.addConstr(power[dev, d, t, "charge"] <= bounds["Battery", "charge"] * cap_design["Battery"])
            model.addConstr(power[dev, d, t, "discharge"] <= bounds["Battery", "discharge"] * cap_design["Battery"])


            # The way it was
            '''
            model.addConstr(storage[dev, d, t] == battery_previous +
                            dt * (power[dev, d, t, "charge"] - power[dev, d, t, "discharge"]))
            model.addConstr(storage[dev, d, t] >= sum(
                                                (1 - bounds["Battery", "DOD"][i]) * cap["Battery"][i] * x[dev, i]
                                                for i in range(number_devices[dev])))
            model.addConstr(storage[dev, d, t] <= sum(cap["Battery"][i] * x[dev, i]
                                                      for i in range(number_devices[dev])))
            model.addConstr(power[dev, d, t, "charge"] ==
                            sum(eta["Battery", "charge"][i] * power[dev, d, t, i, "charge"]
                                for i in range(number_devices[dev])))
            model.addConstr(power[dev, d, t, "discharge"] ==
                            sum(1 / eta["Battery", "discharge"][i] * power[dev, d, t, i, "discharge"]
                                for i in range(number_devices[dev])))

            for i in range(number_devices[dev]):
                model.addConstr(power[dev, d, t, i, "charge"] <= x[dev, i] * bounds["Battery", "charge"][i])
                model.addConstr(power[dev, d, t, i, "discharge"] <= x[dev, i] * bounds["Battery", "discharge"][i])
            '''

            # Electrical heater
            # Heat output
            model.addConstr(heat["EH", d, t] == power["EH", d, t] * eta["EH"])
            # Design restriction
            model.addConstr(cap_design["EH"] >= power["EH", d, t])

            # PV unit
            # Power output
            model.addConstr(power["PV", d, t] <= area["PV"] * eta["PV"] * solar_rad[d, t])
            model.addConstr(power["PV", d, t] == power["PV", d, t, "use"] + power["PV", d, t, "sell"])
            model.addConstr(cap_design["PV"] >= power["PV", d, t])
            # Limit sold power if battery is installed:
            model.addConstr(power["PV", d, t, "sell"] <= cap_design["PV"] -
                            (1-sub_battery["PV", "sell"]) * limit_pv_battery)

            # Solar thermal heater
            model.addConstr(heat["STC", d, t] <= area["STC"] * solar_rad[d, t] * eta["STC"][d, t])
            model.addConstr(power["STC", d, t] == area["STC"] * e_STC[d, t])

            # Heat pump
            # Heat output
            # Linearize: sum(y["HP",d,t,i] for i) * cap_design["TES"] = s_max["TES",d,t]

            # After continuous HP

            model.addConstr(s_max["TES", d, t] <= y["HP", d, t] * bounds["TES", "up"])
            model.addConstr(cap_design["TES"] - s_max["TES", d, t] >= 0)
            model.addConstr(cap_design["TES"] - s_max["TES", d, t] <= (1 - y["HP", d, t]) * bounds["TES", "up"])

            model.addConstr(storage["TES", d, t] <= rho_water * cp_water * (deltaT["HP"] - deltaT["TES"])
                            * s_max["TES", d, t] + rho_water * cp_water * deltaT["TES"] * cap_design["TES"])

            model.addConstr(q_max["HP", d, t] == cop[d, t] / cop_ref * cap_design["HP"])
            model.addConstr(cap_design["HP"] <= bounds["HP", "up"] * y["HP", d, t])
            model.addConstr(heat["HP", d, t] <= q_max["HP", d, t])
            model.addConstr(heat["HP", d, t] >= q_max["HP", d, t] * part_load["HP"])
            model.addConstr(power["HP", d, t] == heat["HP", d, t] / cop[d, t])
            model.addConstr(y["HP", d, t] <= x["HP"])

            # Before continuous HP
            '''
            model.addConstr(s_max["TES", d, t] <= sum(y["HP", d, t, i]
                                                      for i in range(number_devices["HP"])) * bounds["TES", "up"])
            model.addConstr(cap_design["TES"] - s_max["TES", d, t] >= 0)
            model.addConstr(cap_design["TES"] - s_max["TES", d, t] <= (1-sum(y["HP", d, t, i]
                                                                             for i in range(number_devices["HP"]))) * bounds ["TES", "up"])

            model.addConstr(storage["TES", d, t] <= rho_water * cp_water * (deltaT["HP"] - deltaT["TES"]) * s_max["TES",d, t] +
                            rho_water * cp_water * deltaT["TES"] * cap_design["TES"])

            model.addConstr(q_max["HP", d, t] == cop[d, t] / cop_ref * sum(y["HP", d, t, i] * cap["HP"][i]
                                                                           for i in range(number_devices["HP"])))
            model.addConstr(heat["HP", d, t] <= q_max["HP", d, t])
            model.addConstr(heat["HP", d, t] >= q_max["HP", d, t] * part_load["HP"])
            model.addConstr(power["HP", d, t] == heat["HP", d, t] / cop[d, t])

            for i in range(number_devices["HP"]):
                model.addConstr(y["HP", d, t, i] <= x["HP", i])
            '''

            # CHP unit



            # After continuous CHP

            model.addConstr(heat["CHP", d, t] == sum(q_max["CHP", d, t, i] for i in range(number_devices["CHP"])))
            model.addConstr(power["CHP", d, t] == sum(p_max["CHP", d, t, i] for i in range(number_devices["CHP"])))
            model.addConstr(power["CHP", d, t] == power["CHP", d, t, "use"] + power["CHP", d, t, "sell"])
            model.addConstr(energy["CHP", d, t] == 1.0 / eta["CHP", "total"] * (heat["CHP", d, t] + power["CHP", d, t]))
            for i in range(number_devices["CHP"]):
                model.addConstr(y["CHP", d, t, i] <= x["CHP", i])
            # if cap["CHP"][i] < 5:  # According to Spieker and Tsatsaronis
            model.addConstr(q_max["CHP", d, t, 0] == cap_design["CHP0"] / sigma)
            model.addConstr(p_max["CHP", d, t, 0] == cap_design["CHP0"])
            # model.addConstr(cap_design["CHP0"] <= cap["CHP"][0] * y["CHP", d, t, 0])
            # "HAD THESE HERE AT SOME POINT BUT I AM PRETTY SURE THEY ARE WRONG + Already have a similar condition but
            # of course using x and not y (Line 1360)!!!" Now i understand why i put it it is because i was linearizing
            # the equation that had y times cap design but no need since i already have y less than x and x and cap
            # design are related in 1360 This is actually why it never selected a CHP ... because at soon as it had to
            # be off at any point in time ... it meant it was not selected

            # else:
            model.addConstr(p_max["CHP", d, t, 1] <= cap_design["CHP1"])
            model.addConstr(p_max["CHP", d, t, 1] >= cap_design["CHP1"] * part_load["CHP"])
            # model.addConstr(cap_design["CHP1"] <= cap["CHP"][1] * y["CHP", d, t, 1])
            # "HAD THESE HERE AT SOME POINT BUT I AM PRETTY SURE THEY ARE WRONG + Already have a similar condition but
            # of course using x and not y (Line 1360)!!!" Now i understand why i put it it is because i was linearizing
            # the equation that had y times cap design but no need since i already have y less than x and x and cap
            # design are related in 1360 This is actually why it never selected a CHP ... because at soon as it had to
            # be off at any point in time ... it meant it was not selected
            model.addConstr(q_max["CHP", d, t, 1] == 1.0 / params["CHP", "beta"]
                            * (p_max["CHP", d, t, 1] - (params["CHP", "alpha"] * q_CHP[1] + params["CHP", "gamma"] * x["CHP",i])))
            # here i multiplied params gamma by x chp in order to make the product 0 if cap design is 0
            # otherwise qmax will b positive even if cap design is 0
            # and the equation is still linear because params gamma is a constant


            # Before continuous CHP
            '''
            model.addConstr(heat["CHP", d, t] == sum(q_max["CHP", d, t, i] for i in range(number_devices["CHP"])))
            model.addConstr(power["CHP", d, t] == sum(p_max["CHP", d, t, i] for i in range(number_devices["CHP"])))
            model.addConstr(power["CHP", d, t] == power["CHP", d, t, "use"] + power["CHP", d, t, "sell"])
            model.addConstr(energy["CHP", d, t] == 1.0 / eta["CHP", "total"] * (heat["CHP", d, t] + power["CHP", d, t]))
            for i in range(number_devices["CHP"]):
                model.addConstr(y["CHP", d, t, i] <= x["CHP", i])
                if cap["CHP"][i] < 5:  # According to Spieker and Tsatsaronis
                    model.addConstr(q_max["CHP", d, t, i] == y["CHP", d, t, i] * cap["CHP"][i] / sigma)
                    model.addConstr(p_max["CHP", d, t, i] == y["CHP", d, t, i] * cap["CHP"][i])
                else:
                    model.addConstr(p_max["CHP", d, t, i] <= y["CHP", d, t, i] * cap["CHP"][i])
                    model.addConstr(p_max["CHP", d, t, i] >= y["CHP", d, t, i] * cap["CHP"][i] * part_load["CHP"])
                    model.addConstr(q_max["CHP", d, t, i] == 1.0/params["CHP", "beta"] *
                                    (p_max["CHP", d, t, i] - y["CHP", d, t, i] * (params["CHP", "alpha"] * q_CHP[i] + params["CHP", "gamma"])))

            '''
            # Boiler

            # After continuous Boiler
            model.addConstr(heat["Boiler", d, t] == sum(q_max["Boiler", d, t, i] for i in range(number_devices["Boiler"])))
            model.addConstr(energy["Boiler", d, t] == sum(1.0 / eta["Boiler"][i] * q_max["Boiler", d, t, i]
                                                          for i in range(number_devices["Boiler"])))

            for i in range(number_devices["Boiler"]):
                model.addConstr(y["Boiler", d, t, i] <= x["Boiler", i])

            model.addConstr(q_max["Boiler", d, t, 0] <= y["Boiler", d, t, 0] * cap["Boiler"][0])
            model.addConstr(q_max["Boiler", d, t, 1] <= cap_design["Boiler_Conti"])
            model.addConstr(q_max["Boiler", d, t, 0] >= y["Boiler", d, t, 0] * cap["Boiler"][0]
                            * part_load["Boiler"][0])
            model.addConstr(q_max["Boiler", d, t, 1] >= cap_design["Boiler_Conti"]
                            * part_load["Boiler"][1])
            # model.addConstr(cap_design["Boiler_Conti"] <= cap["Boiler"][1] * y["Boiler", d, t, 1])
            # no Need for this since it is actually forcing the "No boiler" decison as soon as the boiler is
            # not on ... the condition to limit cap desing is already put somewhere else and related to x not to y
            # (as it should be)

            # Before continuous Boiler
            '''
            model.addConstr(heat["Boiler", d, t] == sum(q_max["Boiler", d, t, i] for i in range(number_devices["Boiler"])))
            model.addConstr(energy["Boiler", d, t] == sum(1.0 / eta["Boiler"][i] * q_max["Boiler", d, t, i]
                                                          for i in range(number_devices["Boiler"])))

            for i in range(number_devices["Boiler"]):
                model.addConstr(y["Boiler", d, t, i] <= x["Boiler", i])

                model.addConstr(q_max["Boiler", d, t, i] <= y["Boiler", d, t, i] * cap["Boiler"][i])
                model.addConstr(q_max["Boiler", d, t, i] >= y["Boiler", d, t, i] * cap["Boiler"][i] * part_load["Boiler"][i])
            '''

    # ### HUGE CHANGE old End###
    ###
    # 5R1C ISO 13790
    ###############
    # Constraints
    # Decision Variables: only one type activated
    model.addConstr(sum(x["window", i] for i in range(len(U["window"]))) == 1, "window")
    
    for drct2 in direction2: 
        model.addConstr(sum(x["opaque", drct2, j] for j in range(len(U["opaque", drct2]))) 
                        == 1, "opaque_" + drct2)
            
    model.addConstr(sum(x["class", l] for l in range(len(f_class["Am"]))) == 1, "class")

    print('Look Here Here Here')
    print('T_m_o')
    print(T_m_o)
    print('T_m_u')
    print(T_m_u)
    print('T_s_o')
    print(T_s_o)
    print('T_s_u')
    print(T_s_u)



    # 2.4.2) linearizations
    for d in range(number_days):
        for t in range(timesteps):
            timetag = str(d) + "_" + str(t)
            for drct2 in direction2:
                for j in range(len(U["opaque", drct2])):
                    # z_2[t, drct2, j] = x["opaque", drtc2, j] * T_m
                    model.addConstr(z_2[d, t, drct2, j] <= x["opaque", drct2, j] * T_m_o,
                                    "lin1_z_2_"+timetag+"_"+drct2+"_"+str(j))
                    model.addConstr(z_2[d, t, drct2, j] >= x["opaque", drct2, j] * T_m_u,
                                    "lin2_z_2_"+timetag+"_"+drct2+"_"+str(j))
                    model.addConstr(T_m[d, t] - z_2[d, t, drct2, j] <= (1 - x["opaque", drct2, j]) * T_m_o,
                                    "lin3_z_2_"+timetag+"_"+drct2+"_"+str(j))
                    model.addConstr(T_m[d, t] - z_2[d, t, drct2, j] >= (1 - x["opaque", drct2, j]) * T_m_u,
                                    "lin4_z_2_"+timetag+"_"+drct2+"_"+str(j))
            for i in range(len(U["window"])):
                # z_3[t, i] = x["window", i] * T_s
                model.addConstr(z_3[d, t, i] <= x["window", i] * T_s_o,
                                "lin1_z_3_"+timetag+"_"+str(i))
                model.addConstr(z_3[d, t, i] >= x["window", i] * T_s_u,
                                "lin2_z_3_"+timetag+"_"+str(i))
                model.addConstr(T_s[d, t] - z_3[d, t, i] <= (1 - x["window", i]) * T_s_o,
                                "lin3_z_3_"+timetag+"_"+str(i))
                model.addConstr(T_s[d, t] - z_3[d, t, i] >= (1 - x["window", i]) * T_s_u,
                                "lin4_z_3_"+timetag+"_"+str(i))
            for l in range(len(f_class["Am"])):
                # z_6[t, l] = x["class", l] * T_m
                model.addConstr(z_6[d, t, l] <= x["class", l] * T_m_o)
                model.addConstr(z_6[d, t, l] >= x["class", l] * T_m_u)
                model.addConstr(T_m[d, t] - z_6[d, t, l] <= (1 - x["class", l]) * T_m_o)
                model.addConstr(T_m[d, t] - z_6[d, t, l] >= (1 - x["class", l]) * T_m_u)
        # z_7[t, l] = x["class", l] * T_s
                model.addConstr(z_7[d, t, l] <= x["class", l] * T_s_o)
                model.addConstr(z_7[d, t, l] >= x["class", l] * T_s_u)
                model.addConstr(T_s[d, t] - z_7[d, t, l] <= (1 - x["class", l]) * T_s_o)
                model.addConstr(T_s[d, t] - z_7[d, t, l] >= (1 - x["class", l]) * T_s_u)
            
    for drct2 in direction2:
        for j in range(len(U["opaque", drct2])):
            for i in range(len(U["window"])):
                # z_1[drct2, j, i] = x["window", i]*x["opaque", drct2, j]
                model.addConstr(z_1[drct2, j, i] <= x["window", i],
                                "lin1_z1_" + drct2 + "_" + str(j) + "_" + str(i))
                model.addConstr(z_1[drct2, j, i] <= x["opaque", drct2, j], 
                                "lin2_z1_" + drct2 + "_" + str(j) + "_" + str(i))
                model.addConstr(z_1[drct2, j, i] >= x["window", i] + x["opaque", drct2, j] - 1,
                                "lin3_z1_" + drct2 + "_" + str(j) + "_" + str(i))
            for l in range(len(f_class["Am"])):
                # z_4[drct2, j, l] = x["class", l] * x["opaque", drct2, j]
                model.addConstr(z_4[drct2, j, l] <= x["class", l])
                model.addConstr(z_4[drct2, j, l] <= x["opaque", drct2, j])
                model.addConstr(z_4[drct2, j, l] >= x["class", l] + x["opaque", drct2, j] - 1)
    
    for i in range(len(U["window"])):
        for l in range(len(f_class["Am"])):
            # z_5[i, l] = x["class", l] * x["window", i]
            model.addConstr(z_5[i, l] <= x["class", l])
            model.addConstr(z_5[i, l] <= x["window", i])
            model.addConstr(z_5[i, l] >= x["class", l] + x["window", i] - 1)
    
    # A_m (DIN EN ISO 13790, section 12.3.1.2, page 81, table 12)
    A_m = sum(x["class", l] * f_class["Am"][l] * A["f"] for l in range(len(f_class["Am"])))
    
    # C_m (DIN EN ISO 13790, section 12.3.1.2, page 81, table 12)
    C_m = (sum(sum(x["opaque", "wall", j] * kappa["opaque", dev][j] * A["opaque", dev] 
                   for dev in ("wall", "intWall", "ceiling", "intFloor"))
               for j in range(len(kappa["opaque", "wall"]))) + 
           sum(sum(x["opaque", drct4, j] * kappa["opaque", drct4][j] * A["opaque", drct4]
                   for j in range(len(kappa["opaque", drct4]))) 
               for drct4 in direction4))
    
    # Selection of building class based on C_m (scaled to kWh/K instead of J/K)
    model.addConstr(1.0 / 3600000 * C_m / A["f"] <= x["class", 0] * 95000.0 / 3600000
                    + x["class", 1] * 137500.0 / 3600000
                    + x["class", 2] * 212500.0 / 3600000
                    + x["class", 3] * 315000.0 / 3600000
                    + x["class", 4] * 100000000.0 / 3600000)
    model.addConstr(1.0 / 3600000 * C_m / A["f"] >= x["class", 1] * 95000.0 / 3600000
                    + x["class", 2] * 137500.0 / 3600000
                    + x["class", 3] * 212500.0 / 3600000
                    + x["class", 4] * 315000.0 / 3600000)
                                
    # H_tr_w (DIN EN ISO 13790, section 8.3.1, page 44, eq. 18)
    H_tr_w = A["window"] * sum(x["window", i] * U["window"][i] for i in range(len(U["window"])))
    
    # Reference variables to reduce code length
    A_j_k = {}
    B_i_k = {}
    
    for d in range(number_days):
        for t in range(timesteps):
            for drct3 in direction3:
                for j in range(len(U["opaque", "wall"])):
                    A_j_k[d, t, drct3, j] = (U["opaque", "wall"][j]
                                             * R_se["opaque", "wall"][j]
                                             * A["opaque", drct3]
                                             * (alpha_Sc["opaque", "wall"][j]
                                                * I_sol[drct3][d, t] * 1000 - h_r["opaque", "wall"][j]
                                                * F_r[drct3] * Delta_theta_er))
            for drct4 in direction4:
                for j in range(len(U["opaque", drct4])):
                    A_j_k[d, t, drct4, j] = (U["opaque", drct4][j]
                                             * R_se["opaque", drct4][j] * A["opaque", drct4]
                                             * (alpha_Sc["opaque", drct4][j]
                                                * I_sol[drct4][d, t] * 1000 - h_r["opaque", drct4][j]
                                                * F_r[drct4] * Delta_theta_er))
        
            for drct in direction:
                for i in range(len(U["window"])):
                    B_i_k[d, t, drct, i] = A["window", drct] * (g_gl["window"][i]
                                                                * (1-F_F) * I_sol["window", drct][d, t] * 1000
                                                                * F_sh_gl - R_se["window"][i]
                                                                * U["window"][i] * h_r["window"][i]
                                                                * Delta_theta_er * F_r[drct])
    
    phi_sol = {}
    phi_m = {}
    phi_st = {}
    H_tr_em = {}
    for d in range(number_days):
        for t in range(timesteps):
            # heat flow phi_sol [kW]
            # (DIN EN ISO 13790, section 11.3.2, page 67, eq. 43)
            phi_sol[d, t] = (sum(sum(x["opaque", "wall", j] * A_j_k[d, t, drct3, j]
                                     for j in range(len(U["opaque", "wall"])))
                                 for drct3 in direction3) +
                             sum(sum(x["opaque", drct4, j] * A_j_k[d, t, drct4, j]
                                     for j in range(len(U["opaque", drct4])))
                                 for drct4 in direction4) +
                             sum(sum(x["window", i] * B_i_k[d, t, drct, i]
                                     for i in range(len(U["window"])))
                                 for drct in direction))
        
        # heat flow phi_m [kW]
        # (DIN EN ISO 13790, section C2, page 110, eq. C.2)
            phi_m[d, t] = (A_m / A_tot * 0.5 * phi_int[t] +
                           1.0 / A_tot * A["f"] *
                           sum(sum(sum(z_4["wall", j, l] * f_class["Am"][l] * A_j_k[d, t, drct3, j]
                                       for j in range(len(U["opaque", "wall"])))
                                   for drct3 in direction3) +
                               sum(sum(z_4[drct4, j, l] * f_class["Am"][l] * A_j_k[d, t, drct4, j]
                                       for j in range(len(U["opaque", drct4])))
                                   for drct4 in direction4) +
                               sum(sum(z_5[i, l] * f_class["Am"][l] * B_i_k[d, t, drct, i]
                                       for i in range(len(U["window"])))
                                   for drct in direction)
                               for l in range(len(f_class["Am"]))))
        
        # heat flow phi_st [kW] 
        # (DIN EN ISO 13790, section C2, page 110, eq. C.3)  
            phi_st[d, t] = (0.5 * phi_int[t] + phi_sol[d, t] - phi_m[d, t] -
                            H_tr_w / 9.1 / A_tot * 0.5 * phi_int[t] -
                            1.0 / 9.1 / A_tot * A["window"] *
                            sum(sum(sum(z_1["wall", j, i] * U["window"][i] * A_j_k[d, t, drct3, j]
                                        for j in range(len(U["opaque", "wall"])))
                                    for drct3 in direction3) +
                                sum(sum(z_1[drct4, j, i] * U["window"][i] * A_j_k[d, t, drct4, j]
                                        for j in range(len(U["opaque", drct4])))
                                    for drct4 in direction4) +
                                sum(x["window", i] * U["window"][i] * B_i_k[d, t, drct, i]
                                    for drct in direction)
                                for i in range(len(U["window"]))))
            
        # thermal transmittance coefficient H_tr_em [W/K]
        # Simplification: H_tr_em = H_tr_op
        # (DIN EN ISO 13790, section 8.3, page 43)
            H_tr_em[d, t] = sum(A["opaque", drct2] *
                                sum(U["opaque", drct2][j] * x["opaque", drct2, j] * b_tr[drct2][d, t]
                                    for j in range(len(U["opaque", drct2])))
                                for drct2 in direction2)
    
    for d in range(number_days):
        for t in range(timesteps):
            # reference variables to reduce code length
            H_tr_em_x_T_m = sum(A["opaque", drct2] * 
                                sum(z_2[d, t, drct2, j] * U["opaque", drct2][j] * b_tr[drct2][d, t]
                                    for j in range(len(U["opaque", drct2])))
                                for drct2 in direction2)
            
            H_tr_w_x_T_s = A["window"] * sum(z_3[d, t, i] * U["window"][i] for i in range(len(U["window"])))
        
            C_m_x_T_m = (sum(sum(z_2[d, t, "wall", j] * kappa["opaque", dev][j] * A["opaque", dev]
                                 for dev in ("wall", "intWall", "ceiling", "intFloor"))
                             for j in range(len(kappa["opaque", "wall"]))) + 
                         sum(sum(z_2[d, t, drct4, j] * kappa["opaque", drct4][j] * A["opaque", drct4]
                                 for j in range(len(kappa["opaque", drct4])))
                             for drct4 in direction4))
        
            H_tr_ms_x_T_m = sum(z_6[d, t, l] * f_class["Am"][l] * A["f"] * h_ms
                                for l in range(len(f_class["Am"])))
            
            H_tr_ms_x_T_s = sum(z_7[d, t, l] * f_class["Am"][l] * A["f"] * h_ms
                                for l in range(len(f_class["Am"])))
            
        # T_previous for heat transfer of capacity C =  C_m * (T_m[t] * T_m[t-1]) / dt        
            if t == 0:
                C_m_x_T_prev = C_m * T_init  # Initialization
            else:
                C_m_x_T_prev = (sum(sum(z_2[d, t-1, "wall", j] * kappa["opaque", dev][j] * A["opaque", dev]
                                        for dev in ("wall", "intWall", "ceiling", "intFloor")) 
                                    for j in range(len(kappa["opaque", "wall"]))) +
                                sum(sum(z_2[d, t-1, drct4, j] * kappa["opaque", drct4][j] * A["opaque", drct4]
                                        for j in range(len(kappa["opaque", drct4])))
                                    for drct4 in direction4))
               
        ###
        # Linear system of equations to determine T_i, T_s, T_m, Q_HC (in kW)
        # compare (Michalak - 2014 - The simple hourly method of EN ISO 13790) 
        # Using T_i instead of T_op for temperature control


            model.addConstr(H_tr_em_x_T_m + H_tr_ms_x_T_m + C_m_x_T_m / (3600*dt) - H_tr_ms_x_T_s == 
                            phi_m[d, t] + H_tr_em[d, t] * T_e[d, t] + C_m_x_T_prev / (3600*dt),
                            "Ax=b_Gl(1)_"+str(d)+"_"+str(t))
        
            model.addConstr(-H_tr_ms_x_T_m + H_tr_ms_x_T_s + H_tr_is * T_s[d, t] + H_tr_w_x_T_s - H_tr_is * T_i[d, t] ==
                            phi_st[d, t] + H_tr_w * T_e[d, t],
                            "Ax=b_Gl(2)_"+str(d)+"_"+str(t))
        
            model.addConstr(-H_tr_is * T_s[d, t] + H_ve * T_i[d, t] + H_tr_is * T_i[d, t] - 1000 * Q_HC[d, t] ==
                            phi_ia[t] + H_ve * T_e[d, t],
                            "Ax=b_Gl(3)_"+str(d)+"_"+str(t))
        
            model.addConstr(T_i[d, t] >= T_set_min, "Ax=b_Gl(4)_"+str(d)+"_"+str(t))

    print('Look Here Here Here')
    print('H_tr_em_x_T_m')
    print(H_tr_em_x_T_m)
    print('H_tr_ms_x_T_m')
    print(H_tr_ms_x_T_m)
    print('C_m_x_T_m')
    print(C_m_x_T_m)
    print('H_tr_ms_x_T_s')
    print(H_tr_ms_x_T_s)
    print('phi_m[d, t]')
    print(phi_m[d, t])
    print('H_tr_em[d, t]')
    print(H_tr_em[d, t])
    print('T_e[d, t]')
    print(T_e[d, t])
    print('T_e.shape')
    print(T_e.shape)
    print('T_s[d, t]')
    print(T_s[d, t])
    print('T_s.shape')
    print(type(T_s))
    print("(",str(len(T_s)),")")
    print('T_i[d, t]')
    print(T_i[d, t])
    print(type(T_i))
    print('T_i.shape')
    print("(",str(len(T_i)),")")
    print('phi_m[d, t]')
    print(phi_m[d, t])
    print(type(phi_m))
    print('phi_m.shape')
    print("(",str(len(phi_m)),")")
    print('phi_st[d, t]')
    print(phi_st[d, t])
    print(type(phi_st))
    print('phi_st.shape')
    print("(",str(len(phi_st)),")")
    print('phi_ia_[d, t]')
    print(phi_ia[t])
    print(type(phi_ia))
    print('phi_ia.shape')
    print("(",str(len(phi_ia)),")")
    print('A["f"]')
    print(A["f"])
    print('f_class["Am"')
    print(f_class["Am"])
    print('h_ms')
    print(h_ms)



    print('C_m_x_T_prev')
    print(C_m_x_T_prev)
    
    # 5R1C ISO 13790 End
    ###################

    # Prevent boiler+EH and CHP+EH systems
    # After continuous CHP
    model.addConstr(x["EH"] + sum(x["Boiler", i] for i in range(number_devices["Boiler"])) <= 1, "Boiler_or_EH")
    model.addConstr(x["EH"] + sum(x["CHP", i] for i in range(number_devices["CHP"])) <= 1, "CHP_or_EH")


    ###############################################################################
    # Set start values and branching priority
    if options["load_start_vals"]:
        with open(options["filename_start_vals"], "r") as fin:
            for line in fin:
                line_split = line.replace("\n", "").split("\t")
                (model.getVarByName(line_split[0])).Start = int(line_split[1])
    
    for key in x.keys():  # Branch on investment variables first
        x[key].BranchPriority = 100

    # Scenario specific restrictions
    # PV scenario (large taken from status-quo scenario):

    # My Addition
    # Kontinuierliche Variablen

    if options["pv_scenario"]:
        model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
        model.addConstr(sum(x["HP", i] for i in range(len(cap["HP"]))) == 0)
        model.addConstr(x["Battery"] == 0)
        model.addConstr(x["TES", 0] == 1)
        model.addConstr(x["Boiler", 0] == 1)
        model.addConstr(x["EH"] == 0)
        model.addConstr(x["STC"] == 0)
        model.addConstr(x["opaque", "wall", 0] == 1)
        model.addConstr(x["opaque", "roof", 0] == 1)
        model.addConstr(x["window", 0] == 1)

    # The way it was
    '''
    if options["pv_scenario"]:
        model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
        model.addConstr(sum(x["HP", i] for i in range(len(cap["HP"]))) == 0)
        model.addConstr(sum(x["Battery", i] for i in range(len(cap["Battery"]))) == 0)
        model.addConstr(x["TES", 0] == 1)
        model.addConstr(x["Boiler", 0] == 1)
        model.addConstr(x["EH"] == 0)
        model.addConstr(x["STC"] == 0)
        model.addConstr(x["opaque", "wall", 0] == 1)
        model.addConstr(x["opaque", "roof", 0] == 1)
        model.addConstr(x["window", 0] == 1)
    '''

    # My Addition
    # Kontinuierliche Variablen

    # if options["status_quo_with_bat"]:
    #    model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
    #    model.addConstr(x["CHP", 0] == 0)
    #    #model.addConstr(x["CHP", 1] == 0)
    #    # model.addConstr(x["CHP", 0] == 1)
    #    model.addConstr(x["HP"] == 0)
    #    model.addConstr(x["Battery"] == 1)
    #    # model.addConstr(sum(x["Battery", i] for i in range(len(cap["Battery"]))) == 1)
    #    model.addConstr(x["PV"] == 0)
    #    model.addConstr(x["TES", 0] == 1)
    #    # model.addConstr(x["Boiler", 0] == 1)
    #    # model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 1)
    #    model.addConstr(sum(x["Boiler", i] for i in range(len(cap["Boiler"]))) == 0)
    #    model.addConstr(x["EH"] == 1)
    #    model.addConstr(x["STC"] == 0)
    #    model.addConstr(x["opaque", "wall", 0] == 1)
    #    model.addConstr(x["opaque", "roof", 0] == 1)
    #    model.addConstr(x["window", 0] == 1)


    # The way it was
    '''
    #My Addition
    if options["status_quo_with_bat"]:
        model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
        model.addConstr(sum(x["HP", i] for i in range(len(cap["HP"]))) == 0)
        model.addConstr(x["Battery", 0] == 1)
        model.addConstr(sum(x["Battery", i] for i in range(len(cap["Battery"]))) == 1)
        model.addConstr(x["PV"] == 0)
        model.addConstr(x["TES", 0] == 1)
        model.addConstr(x["Boiler", 0] == 1)
        model.addConstr(x["EH"] == 0)
        model.addConstr(x["STC"] == 0)
        model.addConstr(x["opaque", "wall", 0] == 1)
        model.addConstr(x["opaque", "roof", 0] == 1)
        model.addConstr(x["window", 0] == 1)
        '''

    if options["pv_scenario"]:
        model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
        model.addConstr(sum(x["HP", i] for i in range(len(cap["HP"]))) == 0)
        model.addConstr(sum(x["HP", i] for i in range(len(cap["HP"]))) == 0)
        model.addConstr(sum(x["Battery", i] for i in range(len(cap["Battery"]))) == 0)
        model.addConstr(x["PV"] == 0)
        model.addConstr(x["TES", 0] == 1)
        model.addConstr(x["Boiler", 0] == 1)
        model.addConstr(x["EH"] == 0)
        model.addConstr(x["STC"] == 0)
        model.addConstr(x["opaque", "wall", 0] == 1)
        model.addConstr(x["opaque", "roof", 0] == 1)
        model.addConstr(x["window", 0] == 1)
    else:
    #    # EnEV restrictions
        if options["enev_restrictions"]:
            # Set limits
            limit_envelope_U = {"roof": 0.15,
                                "wall": 0.17,
                                "window": 1.1}

            for component in ("roof", "wall"):
                for i in range(len(U["opaque", component])):
                    if U["opaque", component][i] > limit_envelope_U[component]:
                        model.addConstr(x["opaque", component, i] == 0)
            for i in range(len(U["window"])):
                    if U["window"][i] > limit_envelope_U["window"]:
                        model.addConstr(x["window", i] == 0)
    
    ###

    # My addition
    # My new thesis functions
    #model.addConstr(cap_design["Battery"] + cap_design["TES"] == options["total_storage"])

    if options["status_quo_with_bat"]:
       model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
       model.addConstr(x["HP"] == 0)
       model.addConstr(x["TES", 0] == 1)
       model.addConstr(x["Boiler", 0] == 1)
       model.addConstr(x["EH"] == 0)
       model.addConstr(x["STC"] == 0)
       model.addConstr(x["opaque", "wall", 0] == 1)
       model.addConstr(x["opaque", "roof", 0] == 1)
       model.addConstr(x["window", 0] == 1)
       model.addConstr(x["Battery"] == 1)
       model.addConstr(x["PV"] == 0)
       model.addConstr(cap_design["Battery"] == options["total_storage"])

    #if options["status_quo_with_TES"]:
    #    model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
    #    model.addConstr(x["HP"] == 0)
    #    model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 1)
    #    model.addConstr(x["Boiler", 0] == 1)
    #    model.addConstr(x["EH"] == 0)
    #    model.addConstr(x["STC"] == 0)
    #    model.addConstr(x["opaque", "wall", 0] == 1)
    #    model.addConstr(x["opaque", "roof", 0] == 1)
    #    model.addConstr(x["window", 0] == 1)
    #    model.addConstr(x["Battery"] == 0)
    #    model.addConstr(x["PV"] == 0)
    #    model.addConstr(cap_design["TES"] == options["total_storage"])

    if options["scenario_electrify"]:
        model.addConstr(sum(x["CHP", i] for i in range(len(cap["CHP"]))) == 0)
        model.addConstr(x["HP"] == 1)
        model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 0)
        # model.addConstr(x["Boiler", 0] == 0) "I think this was a mistake because it does not force no boiler ... just not the one for free"
        model.addConstr(sum(x["Boiler", i] for i in range(len(cap["Boiler"]))) == 0)
        model.addConstr(x["EH"] == 1)
        model.addConstr(x["STC"] == 0)
        model.addConstr(x["opaque", "wall", 0] == 1)
        model.addConstr(x["opaque", "roof", 0] == 1)
        model.addConstr(x["window", 0] == 1)
        model.addConstr(x["Battery"] == 1)
        model.addConstr(x["PV"] == 1)
        model.addConstr(cap_design["Battery"] == options["total_storage"])


#
    #if options["Battery_or_not"] == False:
    #    model.addConstr(x["Battery"] == 0)
##
    #if options["TES_or_not"] == True:
    #    model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 1)
#
    #if options["TES_or_not"] == False:
    #    model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 0)



    #if options["Battery_or_not"] == True:
    #    model.addConstr( x["Battery"] == 1)
    #    model.addConstr( cap_design["Battery"]+ cap_design["TES"] == options["total_storage"])
    #    if options["TES_or_not"]:
    #        model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 1)
    #    else:
    #        model.addConstr( x["Battery"] == 0)

    #else:
    #    model.addConstr( x["Battery"] == 0)
    #    model.addConstr( cap_design["Battery"]+ cap_design["TES"] == options["total_storage"])
    #    if options["TES_or_not"]:
    #        model.addConstr(sum(x["TES", i] for i in range(len(cap["TES"]))) == 1)
    #    else:
    #        model.addConstr( x["Battery"] == 0)





    # Gurobi parameters
    model.params.TimeLimit = 3600
    model.params.MIPFocus = 1
    model.params.MIPGap = 0.01

    model.write("OmarModelTest.lp")
    #model.write("sa7ebzoubbiNEw.lp")
    model.optimize()
    #model.write("omarnew.lp")
    
    if model.status == gp.GRB.OPTIMAL or model.status == gp.GRB.TIME_LIMIT:
        res_c_total = c_total.X
        # My Addition
        # Kontinuierliche Variablen
        # After continuous HP
        # After continuous CHP
        # No need for CHP here anymore since now it is in number_devices because i made 2 cap values for it

        res_x = {dev: x[dev].X for dev in ["EH", "PV", "STC", "Battery", "HP"]}

        # Before continuous HP
        # Before continuous CHP
        # The way it was
        '''
        res_x = {dev: x[dev].X for dev in ["EH", "PV", "STC"]}
        '''
        for dev in number_devices.keys():
            res_x[dev] = np.array([x[dev, i].X for i in range(number_devices[dev])])
        
        res_x["floor"] = x["opaque", "floor", 0].X
        res_x["roof"] = np.array([x["opaque", "roof", i].X for i in range(len(U["opaque", "roof"]))])
        res_x["wall"] = np.array([x["opaque", "wall", i].X for i in range(len(U["opaque", "wall"]))])
        res_x["window"] = np.array([x["window", i].X for i in range(len(U["window"]))])
        res_x["class"] = np.array([x["class", i].X for i in range(len(f_class["Am"]))])

        # After continuous HP
        # After continuous CHP

        res_y = {}
        for dev in ("Boiler", "CHP", ):
            for i in range(number_devices[dev]):
                res_y[dev, i] = np.array([[y[dev, d, t, i].X for t in range(timesteps)]
                                          for d in range(number_days)])

        for dev in ("HP", ):
            res_y[dev] = np.array([[y[dev, d, t].X for t in range(timesteps)]
                                   for d in range(number_days)])

        res_c_inv = {dev: c_inv[dev].X for dev in c_inv.keys()}
        res_c_serv = {dev: c_serv[dev].X for dev in c_serv.keys()}
        res_c_dem = {dev: c_dem[dev].X for dev in c_dem.keys()}
        res_c_met = c_met.X
        res_cap_design = {dev: cap_design[dev].X for dev in cap_design.keys()}
        res_cap_design["STC"] = area["STC"].X

        # My Addition
        # Kontinuierliche Variablen
        # After continuous HP
        # After continuous CHP
        '''Not there anymore since no need ... capcity returned because they are in teh cap_design dictionary'''

        # Before continuous HP
        # Before continuous CHP
        # The way it was
        '''
        for dev in ("HP", "CHP", "Battery"):
            chosen_or_not = np.max(res_x[dev])
            chosen_argument = np.argmax(res_x[dev])
            res_cap_design[dev] = chosen_or_not * inv[dev][chosen_argument]
        '''
        res_meter = {dev: meter[dev].X for dev in meter.keys()}
        res_revenue = {dev: revenue[dev].X for dev in revenue.keys()}
        
        res_area = {dev: area[dev].X for dev in area.keys()}
        
        res_subsidy = subsidy.X
        res_limit_pv_battery = limit_pv_battery.X
        
        res_energy = {}
        res_power = {}
        res_heat = {}
        res_storage = {}
        for dev in ("TES", "Battery"):
            res_storage[dev] = np.array([[storage[dev, d, t].X for t in range(timesteps)]
                                         for d in range(number_days)])
        
        for dev in ("CHP", "Boiler"):
            res_energy[dev] = np.array([[energy[dev, d, t].X for t in range(timesteps)]
                                        for d in range(number_days)])
        
        for dev in ("Boiler", "CHP", "HP", "EH", "STC"):
            res_heat[dev] = np.array([[heat[dev, d, t].X for t in range(timesteps)]
                                      for d in range(number_days)])
        
        for dev in ("PV", "CHP", "HP", "EH", "Import", "STC", ):
            res_power[dev] = np.array([[power[dev, d, t].X for t in range(timesteps)]
                                       for d in range(number_days)])
        
        for dev in ("PV", "CHP"):
            for method in ("use", "sell"):
                res_power[dev, method] = np.array([[power[dev, d, t, method].X for t in range(timesteps)]
                                                   for d in range(number_days)])
        # My Addition
        # Kontinuierliche Variablen

        res_power["Battery"] = {}
        for dev in ("Battery",):
            for state in ("charge", "discharge"):
                res_power[dev]["total", state] = np.array([[power[dev, d, t, state].X for t in range(timesteps)]
                                                           for d in range(number_days)])

        # The way it was
        '''
        res_power["Battery"] = {}
        for dev in ("Battery",):
            for state in ("charge", "discharge"):
                for i in range(number_devices[dev]):
                    res_power[dev][i, state] = np.array([[power[dev, d, t, i, state].X for t in range(timesteps)]
                                                         for d in range(number_days)])
                res_power[dev]["total", state] = np.array([[power[dev, d, t, state].X for t in range(timesteps)]
                                                           for d in range(number_days)])
                '''
    
        res_q_max = {"HP": np.array([[q_max["HP", d, t].X for t in range(timesteps)]
                                     for d in range(number_days)])}
        for i in range(number_devices["Boiler"]):
            res_q_max["Boiler", i] = np.array([[q_max["Boiler", d, t, i].X for t in range(timesteps)]
                                              for d in range(number_days)])
        res_p_max = {}
        # After continuous CHP
        for i in range(number_devices["CHP"]):
            res_p_max["CHP", i] = np.array([[p_max["CHP", d, t, i].X for t in range(timesteps)]
                                            for d in range(number_days)])
            res_q_max["CHP", i] = np.array([[q_max["CHP", d, t, i].X for t in range(timesteps)]
                                            for d in range(number_days)])

        # 5R1C ISO 13790
        ############### 
        res_Q_HC = np.array([[Q_HC[d, t].X for t in range(timesteps)] for d in range(number_days)])
        res_T_i = np.array([[T_i[d, t].X for t in range(timesteps)] for d in range(number_days)])
        
        A_U = {}
        A_kappa = {}
        A_inv = {}
        print("Envelope:")
        for drct2 in direction2:
            for i in range(len(U["opaque", drct2])):
                if x["opaque", drct2, i].X == 1:
                    print(drct2, ":", i)
                    A_U["opaque", drct2] = A["opaque", drct2] * U["opaque", drct2][i]
                    A_kappa["opaque", drct2] = A["opaque", drct2] * kappa["opaque", drct2][i]
                    A_inv["opaque", drct2] = A["opaque", drct2] * inv["opaque", drct2][i]
        for i in range(len(U["window"])):
            if x["window", i].X == 1:
                print("window:", i)
                A_U["window"] = A["window"] * U["window"][i]
                A_inv["window"] = A["window"] * inv["window"][i]
        
        print("Energy system:")
        for dev in number_devices.keys():
            for i in range(number_devices[dev]):
                if x[dev, i].X == 1:
                    print(dev, "=", cap[dev][i], "[kW] | m3")

        # My Addition
        # Kontinuierliche Variablen
        # After continuous HP
        # After continuous CHP

        for dev in ["EH", "Battery", "HP", ]:
            if x[dev].X == 1:
                print(dev, "=", cap_design[dev].X, u"[kW] | kWh ")

        # The way it was
        # Before continuous HP
        # Before continuous CHP

        '''
        for dev in ["EH", ]:
            if x[dev].X == 1:
                print(dev, "=", cap_design[dev].X, u"[kW]")
        '''
        if x["PV"].X == 1:
            print("PV_area", "=", area["PV"].X, "[m2]")
            print("PV_cap", "=", cap_design["PV"].X, "[kW]")
        if x["STC"].X == 1:
            print("STC", "=", area["STC"].X, "[m2]")
        
        Q_HC_sum = 0    
        for d in range(number_days):
            for t in range(timesteps):
                Q_HC_sum += res_Q_HC[d, t] * weight_days[d]
        print("Q_HC_sum =", Q_HC_sum, "[kWh]")
        
        Q_nHC = ((A_U["window"] + A_U["opaque", "wall"] + A_U["opaque", "roof"] + 
                  A_U["opaque", "floor"] * f_g1 * f_g2 * G_w + 
                  ventilationRate * c_p_air * rho_air * V / 3600) * (T_set_min-T_ne))
        print("DesignHeatLoad =", Q_nHC, "[kW]")
        
        emission_total = emission.X / 1000
        print("Emissions =", emission_total, "t CO2 per year")
        
        print("Costs =", res_c_total, "Euro")


        # After continuous TES
        # After continuous Boiler
        # After continuous CHP
        for dev in ("TES", "Boiler", "CHP"):
            for i in range(number_devices[dev]):
                if x[dev, i].X == 1:
                    print(dev, "=", cap_design[dev].X, "[KW]")

        # Before continuous TES
        # Before continuous Boiler
        # Before continuous CHP
        '''It did not exist!'''


        emissions_of_gas = emissions_gas1.X / 1000
        emissions_of_pv = emissions_PV1.X / 1000
        emissions_of_grid = emissions_grid1.X / 1000
        emissions_of_lca = emissions_lca1.X / 1000
        print("Emissions GAS =", emissions_of_gas, "t CO2 per year")
        print("Emissions PV =", emissions_of_pv, "t CO2 per year")
        print("Emissions GRID =", emissions_of_grid, "t CO2 per year")
        print("Emissions LCA =", emissions_of_lca, "t CO2 per year")
        print("Power Battery Charge =", res_power["Battery"]["total", "charge"], " kw")
        print("Power Battery DisCharge =", res_power["Battery"]["total", "discharge"], " kw")
        # print("Energy Boiler = ", res_energy["Boiler"], "t CO2 per year")
        # print("Energy CHP = ", res_energy["CHP"], "t CO2 per year")
        # print("x CHP 0 =  " + str(res_x["CHP"][0]))
        # print(" Cap_design_CHP 0 =  " + str(res_cap_design["CHP0"]))
        # print(" q_max_CHP 0 =  " + str(res_q_max["CHP", 0]))
        # print(" p_max_CHP 0 =  " + str(res_p_max["CHP", 0]))
        # print("x CHP 1 =  " + str(res_x["CHP"][1]))
        # print(" Cap_design_CHP 1 =  " + str(res_cap_design["CHP1"]))
        # print(" q_max_CHP 1 =  " + str(res_q_max["CHP", 1]))
        # print(" p_max_CHP 1 =  " + str(res_p_max["CHP", 1]))





        #if x["TES", 0].X == 1:
        #    print("TES CAPACITY", "=", cap_design["TES"].X, "[kW]")
        #if x["TES", 1].X == 1:
        #    print("TES CAPACITY", "=", cap_design["TES"].X, "[kW]")

        # Before continuous TES
        '''Did not exist'''

        # My Addition
        Cap_Batt = cap_design["Battery"].X
        Cap_TES = cap_design["TES"].X

        # 5R1C ISO 13790 End
        ###################
    else:  # Error in optimization
        model.computeIIS()
        model.write("model.ilp")
        print('\nConstraints:')
        for c in model.getConstrs():
            if c.IISConstr:
                print('%s' % c.constrName)
        print('\nBounds:')
        for v in model.getVars():
            if v.IISLB > 0:
                print('Lower bound: %s' % v.VarName)
            elif v.IISUB > 0:
                print('Upper bound: %s' % v.VarName)

    # Save results
    with open(options["filename_results"], "wb") as fout:
        pickle.dump(res_c_total, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(model.MIPGap, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(model.Params.TimeLimit, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_x, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_y, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_c_inv, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_c_serv, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_c_dem, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_c_met, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_cap_design, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_meter, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_revenue, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_area, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_subsidy, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_limit_pv_battery, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_energy, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_power, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_heat, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_storage, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_q_max, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_p_max, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_Q_HC, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(res_T_i, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(A_U, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(A_kappa, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(A_inv, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(Q_HC_sum, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(Q_nHC, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(emission_total, fout, pickle.HIGHEST_PROTOCOL)
        pickle.dump(emission_max, fout, pickle.HIGHEST_PROTOCOL)

    if options["store_start_vals"]:
        with open(options["filename_start_vals"], "w") as fout:
            for var in model.getVars():
                if var.VType == "B":
                    fout.write(var.VarName + "\t" + str(int(var.X)) + "\n")


    # For my new run multi obj
    return(res_c_total, emission_total, emissions_of_gas, emissions_of_pv, emissions_of_grid, emissions_of_lca,
           res_x, res_cap_design, res_power,res_storage["Battery"], clustered)

    # For my CO2_vs_Storage
    # return (res_c_total, emission_total, Cap_Batt, Cap_TES)


    #For my extra Function
    #return (res_c_total, emission_total, Cap_Batt,
    #         res_power["Import"], res_power["Battery"]["total", "charge"], res_power["Battery"]["total", "discharge"],
    #         res_storage["Battery"]
    #         )

    # How it was
    #return (res_c_total, emission_total)
